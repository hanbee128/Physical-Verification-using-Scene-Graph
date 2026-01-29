#!/usr/bin/env python3
"""
Evaluation script for Physical Guard results.
This script reads the generated plans from `physical_guard_set3_result_*.txt`,
executes them using AI2-THOR, and compares the final state with the expected results
defined in `data/final_test/FloorPlan1.json`.
"""

import os
import sys
import json
import glob
import re
import argparse
import time
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple

# Add the scripts directory to sys.path to import modules
current_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(current_dir)

# Return to AI2ThorExecutor
from ai2thor_connector_ithor import AI2ThorExecutor

# Excel 파일 생성을 위한 라이브러리
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl이 설치되지 않았습니다. 설치 중...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

def parse_json_plan_file(file_path: str) -> Dict[str, str]:
    """
    Parses the JSON plan file to extract tasks and their corresponding programs.
    Returns a dictionary mapping task names to program code.
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        plan_dict = json.load(f)
    
    plans = {}
    for task_name, program_code in plan_dict.items():
        # Normalize task name to lowercase for robust matching
        normalized_task = task_name.strip().lower()
        plans[normalized_task] = program_code
    
    return plans

def parse_plan_file(file_path: str) -> Dict[str, str]:
    """
    Parses the plan file (txt) to extract tasks and their corresponding programs.
    Returns a dictionary mapping task names to program code.
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    plans = {}
    # Split by "Task: "
    parts = content.split("Task: ")
    for part in parts:
        if not part.strip():
            continue
        
        # Extract task name (first line)
        lines = part.strip().split('\n')
        task_name = lines[0].strip()
        
        # Extract code (skip separator line if present)
        code_lines = []
        start_collecting = False
        for line in lines[1:]:
            if line.startswith("==="):
                start_collecting = True
                continue
            if line.startswith("Physical Verification Summary:"): # Stop parsing at summary
                break
            if start_collecting or not line.startswith("==="): # If no separator, start collecting immediately or after separator
                 code_lines.append(line)
        
        # Clean up code
        code = '\n'.join(code_lines).strip()
        
        # If code is empty, try to extract from function definition
        if not code:
            # Look for function definition: def function_name():
            for line in lines:
                if line.strip().startswith("def ") and "():" in line:
                    # Extract task name from function name or comment
                    # Try to find "# Task: " comment
                    for comment_line in lines:
                        if "# Task:" in comment_line:
                            task_name_from_comment = comment_line.split("# Task:")[-1].strip()
                            if task_name_from_comment:
                                task_name = task_name_from_comment
                                break
                    break
        
        if code and task_name:
             # Normalize task name to lowercase for robust matching
             normalized_task = task_name.strip().lower()
             plans[normalized_task] = code
             
    return plans

def load_expected_results(json_path: str) -> List[Dict[str, Any]]:
    """Loads expected object states from JSON file."""
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def verify_object_state(executor: AI2ThorExecutor, expected_state: Dict[str, Any]) -> Tuple[int, int]:
    """
    Verifies object state and returns (passed_count, total_count).
    """
    object_name = expected_state['name']
    expected_contains = expected_state.get('contains', [])
    expected_status = expected_state.get('state', 'None')
    
    passed_count = 0
    total_count = 0
    
    obj_id = executor._find_object_id(object_name)
    if not obj_id:
        print(f"  ❌ Object '{object_name}' not found in simulator.")
        # If object not found, we count the conditions but they result in 0 passed.
        # Count 'contains' items
        total_count += len(expected_contains)
        # Count 'state' if not None
        if expected_status != 'None':
             total_count += 1
        return 0, total_count

    # Get object metadata from last event
    all_objects = executor.controller.last_event.metadata['objects']
    obj_meta = next((o for o in all_objects if o['objectId'] == obj_id), None)
    
    if not obj_meta:
        print(f"  ❌ Metadata for '{object_name}' ({obj_id}) not found.")
        total_count += len(expected_contains)
        if expected_status != 'None':
             total_count += 1
        return 0, total_count

    # Check containment using receptacleObjectIds
    if expected_contains:
        # Get object IDs of things currently inside the target object using receptacleObjectIds
        receptacle_object_ids = obj_meta.get('receptacleObjectIds', [])
        contained_objects = []
        
        # receptacleObjectIds에 있는 objectId로 실제 객체 정보 찾기
        for contained_obj_id in receptacle_object_ids:
            contained_obj = next((o for o in all_objects if o.get('objectId') == contained_obj_id), None)
            if contained_obj:
                contained_objects.append({
                    'objectId': contained_obj_id,
                    'objectType': contained_obj.get('objectType', '')
                })

        # Check if each expected item is present
        for item in expected_contains:
            total_count += 1
            found = False
            found_object_id = None
            
            for contained in contained_objects:
                contained_type = contained.get('objectType', '')
                # objectType으로 매칭 (예: "Egg", "Apple" 등)
                if item.lower() in contained_type.lower() or contained_type.lower() in item.lower():
                    found = True
                    found_object_id = contained.get('objectId')
                    break
            
            if not found:
                # Debug: find where it actually is
                actual_parent = "None"
                for o in all_objects:
                    obj_type = o.get('objectType', '')
                    if item.lower() in obj_type.lower() or obj_type.lower() in item.lower():
                        # parentReceptacle 또는 parentReceptacles 확인
                        pid = o.get('parentReceptacle')
                        if not pid:
                            parent_receptacles = o.get('parentReceptacles', [])
                            if parent_receptacles:
                                pid = parent_receptacles[0]
                        if pid:
                            # parent receptacle의 objectType 찾기
                            parent_obj = next((p for p in all_objects if p.get('objectId') == pid), None)
                            if parent_obj:
                                actual_parent = parent_obj.get('objectType', pid)
                            else:
                                actual_parent = pid
                            break
                             
                print(f"  ❌ Expected '{item}' in '{object_name}', but not found.")
                print(f"     Found in receptacle: {[c['objectType'] for c in contained_objects]}")
                print(f"     (Debug: '{item}' seems to be in '{actual_parent}')")
            else:
                print(f"  ✅ Verified '{item}' is in '{object_name}' (objectId: {found_object_id}).")
                passed_count += 1

    # Check state
    if expected_status != 'None':
        total_count += 1
        state_passed = False
        
        if expected_status == 'Sliced':
            is_sliced = obj_meta.get('isSliced', False)
            if is_sliced:
                print(f"  ✅ Verified '{object_name}' is Sliced.")
                state_passed = True
            else:
                print(f"  ❌ Expected '{object_name}' to be Sliced, but isSliced={is_sliced}.")
                
        elif expected_status == 'Broken':
            is_broken = obj_meta.get('isBroken', False)
            if is_broken:
                print(f"  ✅ Verified '{object_name}' is Broken.")
                state_passed = True
            else:
                print(f"  ❌ Expected '{object_name}' to be Broken, but isBroken={is_broken}.")
                
        elif expected_status == 'Opened':
            is_open = obj_meta.get('isOpen', False)
            if is_open:
                print(f"  ✅ Verified '{object_name}' is Opened (isOpen={is_open}).")
                state_passed = True
            else:
                print(f"  ❌ Expected '{object_name}' to be Opened, but isOpen={is_open}.")
                
        elif expected_status == 'Closed':
            is_open = obj_meta.get('isOpen', False)
            if not is_open:
                print(f"  ✅ Verified '{object_name}' is Closed (isOpen={is_open}).")
                state_passed = True
            else:
                print(f"  ❌ Expected '{object_name}' to be Closed, but isOpen={is_open}.")
                
        elif expected_status == 'ToggledOn' or expected_status == 'On':
            is_toggled = obj_meta.get('isToggled', False)
            if is_toggled:
                print(f"  ✅ Verified '{object_name}' is {'On' if expected_status == 'On' else 'ToggledOn'} (isToggled={is_toggled}).")
                state_passed = True
            else:
                print(f"  ❌ Expected '{object_name}' to be {'On' if expected_status == 'On' else 'ToggledOn'}, but isToggled={is_toggled}.")
                
        elif expected_status == 'ToggledOff' or expected_status == 'Off':
            is_toggled = obj_meta.get('isToggled', False)
            if not is_toggled:
                print(f"  ✅ Verified '{object_name}' is {'Off' if expected_status == 'Off' else 'ToggledOff'} (isToggled={is_toggled}).")
                state_passed = True
            else:
                print(f"  ❌ Expected '{object_name}' to be {'Off' if expected_status == 'Off' else 'ToggledOff'}, but isToggled={is_toggled}.")
        else:
            print(f"  ⚠ Unknown state '{expected_status}' for '{object_name}'.")
            # 알 수 없는 상태는 통과로 처리하지 않음
        
        if state_passed:
            passed_count += 1

    return passed_count, total_count

def execute_and_evaluate_task(
    executor: AI2ThorExecutor,
    task_def: Dict[str, Any],
    program_code: str,
    method_name: str = "Unknown"
) -> Dict[str, Any]:
    """
    Execute a task and evaluate the results.
    Returns evaluation result dictionary.
    """
    task_name = task_def['task']
    print(f"\n[{method_name}] Evaluating Task: {task_name}")
    print("="*50)
    
    # Reset environment for each task to ensure clean state
    executor.controller.reset(executor.scene)
    
    # Initialize agent again
    executor.controller.step(dict(
        action='Initialize',
        agentMode="default",
        snapGrid=False,
        gridSize=0.25,
        visibilityDistance=1.5,
        fieldOfView=120,
        agentCount=1
    ))
    
    # Parse and execute actions
    lines = program_code.split('\n')
    execution_failed = False
    
    # 총 생성된 action 수 계산 (assert 제외)
    total_generated_actions = 0
    for line in lines:
        line = line.strip()
        if not line:
            continue
        if line.startswith("#"):
            continue  # Skip comments
        match = re.match(r"(\w+)\((.*)\)", line)
        if match:
            action = match.group(1)
            if action != "assert":  # assert는 액션이 아님
                total_generated_actions += 1
    
    # Track executed actions for JSON report
    executed_actions_log = []
    
    for line in lines:
        line = line.strip()
        if not line: 
            continue
        if line.startswith("#"):
            continue  # Skip comments
        
        # Simple parsing regex
        match = re.match(r"(\w+)\((.*)\)", line)
        if match:
            action = match.group(1)
            args_str = match.group(2)
            
            # assert 문은 실행하지 않음 (Logical precondition, AI2-THOR 액션이 아님)
            if action == "assert":
                continue
            
            # Split args by comma, ignoring commas inside quotes
            try:
                if args_str.strip():
                    # eval 사용 시 "is" with literal SyntaxWarning 방지: 인자만 추출
                    args = eval(f"({args_str},)")  # Force tuple
                else:
                    args = ()
                
                target = args[0] if len(args) > 0 else None
                receptacle = args[1] if len(args) > 1 else None
                
                print(f"  Running: {line}")
                
                success = False
                
                # snake_case mapping: PickupObject -> pickup_object
                method_name_action = re.sub(r'(?<!^)(?=[A-Z])', '_', action).lower()
                
                if hasattr(executor, method_name_action):
                    func = getattr(executor, method_name_action)
                    try:
                        if receptacle:
                            success = func(target, receptacle)
                        elif target:
                            success = func(target)
                        else:
                            success = func()
                    except Exception as e:
                        print(f"    Error executing {method_name_action}: {e}")
                        success = False
                else:
                    # Fallback for some common ones
                    if action == "GoToObject" and hasattr(executor, "goto_object"):
                         success = executor.goto_object(target, target_distance=0.5, success_distance=1.0)
                    elif action == "PutObject" and hasattr(executor, "put_object"):
                         success = executor.put_object(target, receptacle)
                    elif action == "PickupObject" and hasattr(executor, "pickup_object"):
                         success = executor.pickup_object(target)
                    elif action == "OpenObject" and hasattr(executor, "open_object"):
                         success = executor.open_object(target)
                    elif action == "CloseObject" and hasattr(executor, "close_object"):
                         success = executor.close_object(target)
                    elif action == "SliceObject" and hasattr(executor, "slice_object"):
                         success = executor.slice_object(target)
                    elif action == "BreakObject" and hasattr(executor, "break_object"):
                         success = executor.break_object(target)
                    elif action == "ToggleObjectOn" and hasattr(executor, "toggle_on"):
                         success = executor.toggle_on(target)
                    elif action == "ToggleObjectOff" and hasattr(executor, "toggle_off"):
                         success = executor.toggle_off(target)
                    else:
                         print(f"    Warn: Unknown action method '{method_name_action}' or '{action}'")
                         success = False

                if not success:
                    print(f"    ❌ Action failed: {line}")
                    execution_failed = True
                
                # Log action execution
                last_event = executor.controller.last_event
                last_action_success = last_event.metadata.get('lastActionSuccess', False) if last_event else success
                error_message = last_event.metadata.get('errorMessage', '') if last_event else ''
                if not success and not error_message:
                    error_message = "Action method returned False"

                executed_actions_log.append({
                    "line": line,
                    "action": action,
                    "target_object": target,
                    "receptacle": receptacle,
                    "success": success,
                    "lastActionSuccess": last_action_success,
                    "errorMessage": error_message
                })

            except Exception as e:
                print(f"    Error parsing/executing line '{line}': {e}")
                execution_failed = True
                executed_actions_log.append({
                    "line": line,
                    "error": str(e),
                    "success": False
                })
    
    # Verify final state
    print(f"\n[{method_name}] Verifying object states...")
    print("="*50)
    all_passed = True
    task_passed_conds = 0
    task_total_conds = 0

    # object_states 키가 있는지 확인
    object_states = task_def.get('object_states', [])
    if not object_states:
        print(f"  ⚠️  No object_states defined for task '{task_name}'")
        print(f"    → Skipping verification")
        gcr = 0.0
        task_total_conds = 0
        task_passed_conds = 0
    else:
        for obj_state in object_states:
            print(f"  Checking: {obj_state['name']}")
            p_count, t_count = verify_object_state(executor, obj_state)
            task_passed_conds += p_count
            task_total_conds += t_count
            print(f"    → Passed: {p_count}/{t_count} conditions")
        
        if task_total_conds > 0:
            gcr = (task_passed_conds / task_total_conds) * 100
            if task_passed_conds < task_total_conds:
                all_passed = False
        else:
            gcr = 0.0
            print("    ⚠ No conditions to verify.")
        
    status = "PASSED" if all_passed and not execution_failed else "FAILED"
    if execution_failed:
         status += " (Execution Errors)"
         
    # Exec 계산: 총 생성된 action 수 대비 성공적으로 실행된 action 비율 (실패한 액션은 제외)
    executed_actions_count = len(executed_actions_log)
    successful_actions_count = sum(
        1 for a in executed_actions_log
        if a.get("lastActionSuccess") is True or (a.get("success") is True and a.get("lastActionSuccess") is not False)
    )
    if total_generated_actions > 0:
        exec_value = (successful_actions_count / total_generated_actions) * 100
    else:
        exec_value = 0.0 if successful_actions_count == 0 else 100.0
    
    print(f"\n[{method_name}] Task Result: {status}")
    print(f"[{method_name}] GCR: {gcr:.1f}% ({task_passed_conds}/{task_total_conds} conditions passed)")
    print(f"[{method_name}] Exec: {exec_value:.1f}% ({successful_actions_count}/{total_generated_actions} actions succeeded)")
    
    return {
        "task": task_name,
        "method": method_name,
        "status": status,
        "gcr": round(gcr, 2),
        "exec": round(exec_value, 2),
        "executed_actions_count": executed_actions_count,
        "successful_actions_count": successful_actions_count,
        "total_generated_actions": total_generated_actions,
        "passed_conditions": task_passed_conds,
        "total_conditions": task_total_conds,
        "verification_passed": all_passed,
        "execution_clean": not execution_failed,
        "executed_actions": executed_actions_log
    }

def save_results_to_excel(physical_guard_results: List[Dict], baseline_results: List[Dict], fp_number: str, output_file: str):
    """평가 결과를 Excel 파일로 저장"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    
    # 스타일 정의
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # 헤더 작성
    headers = ["Task", "Physical Guard GCR (%)", "Baseline GCR (%)", "차이 (PG - BL)", "Physical Guard Exec (%)", "Baseline Exec (%)", "Physical Guard Status", "Baseline Status"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # 데이터 작성
    pg_dict = {r['task']: r for r in physical_guard_results}
    bl_dict = {r['task']: r for r in baseline_results}
    
    all_tasks = set(pg_dict.keys()) | set(bl_dict.keys())
    row = 2
    
    total_pg_gcr = 0
    total_bl_gcr = 0
    total_pg_exec = 0
    total_bl_exec = 0
    count_pg = 0
    count_bl = 0
    count_pg_exec = 0
    count_bl_exec = 0
    
    for task_name in sorted(all_tasks):
        pg_res = pg_dict.get(task_name, {})
        bl_res = bl_dict.get(task_name, {})
        
        pg_gcr = pg_res.get('gcr', 0) if 'gcr' in pg_res else None
        bl_gcr = bl_res.get('gcr', 0) if 'gcr' in bl_res else None
        pg_exec = pg_res.get('exec', 0) if 'exec' in pg_res else None
        bl_exec = bl_res.get('exec', 0) if 'exec' in bl_res else None
        
        pg_status = pg_res.get('status', 'N/A')
        bl_status = bl_res.get('status', 'N/A')
        
        diff = (pg_gcr - bl_gcr) if (pg_gcr is not None and bl_gcr is not None) else None
        
        ws.cell(row=row, column=1, value=task_name)
        ws.cell(row=row, column=2, value=round(pg_gcr, 2) if pg_gcr is not None else "N/A")
        ws.cell(row=row, column=3, value=round(bl_gcr, 2) if bl_gcr is not None else "N/A")
        ws.cell(row=row, column=4, value=round(diff, 2) if diff is not None else "N/A")
        ws.cell(row=row, column=5, value=round(pg_exec, 2) if pg_exec is not None else "N/A")
        ws.cell(row=row, column=6, value=round(bl_exec, 2) if bl_exec is not None else "N/A")
        ws.cell(row=row, column=7, value=pg_status)
        ws.cell(row=row, column=8, value=bl_status)
        
        if pg_gcr is not None:
            total_pg_gcr += pg_gcr
            count_pg += 1
        if bl_gcr is not None:
            total_bl_gcr += bl_gcr
            count_bl += 1
        if pg_exec is not None:
            total_pg_exec += pg_exec
            count_pg_exec += 1
        if bl_exec is not None:
            total_bl_exec += bl_exec
            count_bl_exec += 1
        
        row += 1
    
    # 평균 행 추가
    row += 1
    avg_pg_gcr = total_pg_gcr / count_pg if count_pg > 0 else 0
    avg_bl_gcr = total_bl_gcr / count_bl if count_bl > 0 else 0
    avg_pg_exec = total_pg_exec / count_pg_exec if count_pg_exec > 0 else 0
    avg_bl_exec = total_bl_exec / count_bl_exec if count_bl_exec > 0 else 0
    avg_diff = avg_pg_gcr - avg_bl_gcr
    
    ws.cell(row=row, column=1, value="평균")
    ws.cell(row=row, column=2, value=round(avg_pg_gcr, 2))
    ws.cell(row=row, column=3, value=round(avg_bl_gcr, 2))
    ws.cell(row=row, column=4, value=round(avg_diff, 2))
    ws.cell(row=row, column=5, value=round(avg_pg_exec, 2))
    ws.cell(row=row, column=6, value=round(avg_bl_exec, 2))
    
    # 평균 행 스타일
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # 열 너비 조정
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 25
    
    # 기본 시트 제거
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["Sheet"])
    
    # 파일 저장
    wb.save(output_file)
    print(f"\n✅ Excel 파일 저장 완료: {output_file}")

def evaluate_single_fp(fp_number: str, folder_name: str, test_file: str) -> Tuple[List[Dict], List[Dict]]:
    """단일 FP에 대한 평가 실행"""
    fp_folder = f"FP{fp_number}"
    
    # Find JSON files
    folder_physical_guard_pattern = os.path.join(current_dir, "../results", folder_name, fp_folder, "physical_guard_result_*.json")
    folder_baseline_pattern = os.path.join(current_dir, "../results", folder_name, fp_folder, "baseline_result_*.json")
    
    physical_guard_files = glob.glob(folder_physical_guard_pattern)
    baseline_files = glob.glob(folder_baseline_pattern)
    
    # Try relative path if absolute path didn't work
    if not physical_guard_files:
        folder_physical_guard_pattern = os.path.join("results", folder_name, fp_folder, "physical_guard_result_*.json")
        physical_guard_files = glob.glob(folder_physical_guard_pattern)
    
    if not baseline_files:
        folder_baseline_pattern = os.path.join("results", folder_name, fp_folder, "baseline_result_*.json")
        baseline_files = glob.glob(folder_baseline_pattern)
    
    # Fallback: Check parent folder
    if not physical_guard_files:
        search_pattern = os.path.join(current_dir, "../results", folder_name, "physical_guard_result_*.json")
        physical_guard_files = glob.glob(search_pattern)
        if not physical_guard_files:
            search_pattern = os.path.join("results", folder_name, "physical_guard_result_*.json")
            physical_guard_files = glob.glob(search_pattern)
    
    if not baseline_files:
        search_pattern = os.path.join(current_dir, "../results", folder_name, "baseline_result_*.json")
        baseline_files = glob.glob(search_pattern)
        if not baseline_files:
            search_pattern = os.path.join("results", folder_name, "baseline_result_*.json")
            baseline_files = glob.glob(search_pattern)
    
    # Get latest files if multiple found
    physical_guard_file = None
    baseline_file = None
    
    if physical_guard_files:
        physical_guard_file = max(physical_guard_files, key=os.path.getmtime)
    
    if baseline_files:
        baseline_file = max(baseline_files, key=os.path.getmtime)
    
    if not physical_guard_file and not baseline_file:
        print(f"⚠️  FP{fp_number}: 결과 파일을 찾을 수 없습니다.")
        return [], []
    
    # Load plans
    physical_guard_plans = {}
    baseline_plans = {}
    
    if physical_guard_file:
        physical_guard_plans = parse_json_plan_file(physical_guard_file)
    
    if baseline_file:
        baseline_plans = parse_json_plan_file(baseline_file)
    
    # Load expected results
    expected_results = load_expected_results(test_file)
    
    # Initialize Executors
    scene_name = f"FloorPlan{fp_number}"
    physical_guard_exe = None
    baseline_exe = None
    
    if physical_guard_plans:
        physical_guard_exe = AI2ThorExecutor(
            scene=scene_name,
            headless=False,
            save_video=False
        )
        physical_guard_exe.initialize()
    
    if baseline_plans:
        baseline_exe = AI2ThorExecutor(
            scene=scene_name,
            headless=False,
            save_video=False
        )
        baseline_exe.initialize()
    
    # Evaluate tasks
    physical_guard_results = []
    baseline_results = []
    
    def _norm_task(s: str) -> str:
        return (s or "").lower().strip().replace(" ", "")

    for task_def in expected_results:
        task_name = task_def['task']
        task_key = task_name.strip().lower()
        
        # Evaluate Physical Guard
        if physical_guard_plans and physical_guard_exe:
            matched_key = None
            if task_key in physical_guard_plans:
                matched_key = task_key
            else:
                task_n = _norm_task(task_name)
                for plan_key in physical_guard_plans.keys():
                    if _norm_task(plan_key) == task_n:
                        matched_key = plan_key
                        break
                if not matched_key:
                    for plan_key in physical_guard_plans.keys():
                        if task_key in plan_key or plan_key in task_key:
                            matched_key = plan_key
                            break
            
            if matched_key:
                program_code = physical_guard_plans[matched_key]
                result = execute_and_evaluate_task(
                    physical_guard_exe, task_def, program_code, "Physical Guard"
                )
                physical_guard_results.append(result)
            else:
                physical_guard_results.append({
                    "task": task_name,
                    "method": "Physical Guard",
                    "status": "SKIPPED",
                    "reason": "No Plan"
                })
        
        # Evaluate Baseline
        if baseline_plans and baseline_exe:
            matched_key = None
            if task_key in baseline_plans:
                matched_key = task_key
            else:
                task_n = _norm_task(task_name)
                for plan_key in baseline_plans.keys():
                    if _norm_task(plan_key) == task_n:
                        matched_key = plan_key
                        break
                if not matched_key:
                    for plan_key in baseline_plans.keys():
                        if task_key in plan_key or plan_key in task_key:
                            matched_key = plan_key
                            break
            
            if matched_key:
                program_code = baseline_plans[matched_key]
                result = execute_and_evaluate_task(
                    baseline_exe, task_def, program_code, "Baseline"
                )
                baseline_results.append(result)
            else:
                baseline_results.append({
                    "task": task_name,
                    "method": "Baseline",
                    "status": "SKIPPED",
                    "reason": "No Plan"
                })
    
    # Cleanup
    if physical_guard_exe:
        physical_guard_exe.close()
    if baseline_exe:
        baseline_exe.close()
    
    return physical_guard_results, baseline_results

def main():
    parser = argparse.ArgumentParser(description="Evaluate Physical Guard and Baseline Results")
    parser.add_argument("--physical_guard_json", type=str, help="Path to physical_guard_result JSON file (optional)")
    parser.add_argument("--baseline_json", type=str, help="Path to baseline_result JSON file (optional)")
    parser.add_argument("--test_file", type=str, default=None, help="Path to the test definition JSON")
    parser.add_argument("--folder", type=str, default="Kitchen", 
                       choices=["Kitchen", "LivingRoom", "BedRoom", "BathRoom"],
                       help="Folder name in results/ directory to search for JSON files (default: Kitchen)")
    parser.add_argument("--fp-number", type=str, default=None,
                       help="FP number (e.g., 1, 2, 216, 224, 325, 326, 403, 425). If not provided, will process all FPs in folder.")
    parser.add_argument("--all-fps", action="store_true",
                       help="Process all FPs in the selected folder")
    args = parser.parse_args()

    folder_name = args.folder
    
    # 폴더 안의 모든 FP 폴더 찾기
    folder_dir = os.path.join(current_dir, "../results", folder_name)
    if not os.path.exists(folder_dir):
        folder_dir = os.path.join("results", folder_name)
    
    if not os.path.exists(folder_dir):
        print(f"❌ 폴더를 찾을 수 없습니다: {folder_dir}")
        return
    
    # FP 폴더 찾기 (FP{num} 형식)
    fp_folders = []
    if os.path.exists(folder_dir):
        for item in os.listdir(folder_dir):
            item_path = os.path.join(folder_dir, item)
            if os.path.isdir(item_path) and item.startswith("FP"):
                # FP{num} 형식에서 숫자 추출
                match = re.match(r'FP(\d+)', item)
                if match:
                    fp_folders.append((item, match.group(1)))
    
    if not fp_folders:
        print(f"❌ {folder_dir}에 FP 폴더를 찾을 수 없습니다.")
        return
    
    # FP 번호 필터링
    if args.fp_number:
        fp_folders = [(folder, num) for folder, num in fp_folders if num == args.fp_number]
        if not fp_folders:
            print(f"❌ FP{args.fp_number} 폴더를 찾을 수 없습니다.")
            return
    
    print(f"\n📂 {folder_name} 폴더에서 {len(fp_folders)}개의 FP 폴더를 찾았습니다:")
    for folder, num in fp_folders:
        print(f"   - {folder} (FloorPlan{num})")
    
    # 각 FP에 대해 평가 실행
    for folder, fp_number in fp_folders:
        print(f"\n{'='*70}")
        print(f"🔄 FloorPlan{fp_number} 평가 시작")
        print(f"{'='*70}")
        
        # Test file 자동 선택
        test_file = args.test_file
        if not test_file:
            test_file_path = os.path.join("data/final_test", f"FloorPlan{fp_number}.json")
            if os.path.exists(test_file_path):
                test_file = test_file_path
            else:
                print(f"⚠️  Test 파일을 찾을 수 없습니다: {test_file_path}")
                continue
        
            # 평가 실행
        try:
            physical_guard_results, baseline_results = evaluate_single_fp(
                fp_number=fp_number,
                folder_name=folder_name,
                test_file=test_file
            )
            
            if not physical_guard_results and not baseline_results:
                print(f"⚠️  FloorPlan{fp_number}: 평가 결과가 없습니다.")
                continue
            
            # 기존 출력 형식으로 요약 출력
            print("\n" + "="*70)
            print("EVALUATION SUMMARY - COMPARISON")
            print("="*70)
            
            # Physical Guard Summary
            if physical_guard_results:
                print("\n📊 PHYSICAL GUARD RESULTS:")
                print("-" * 70)
                total_gcr = 0
                count_gcr = 0
                for res in physical_guard_results:
                    if 'gcr' in res:
                        status_icon = "✅" if res['status'] == "PASSED" else "❌"
                        print(f"{status_icon} {res['task']}")
                        print(f"   Status: {res['status']}")
                        print(f"   GCR: {res['gcr']:.1f}%")
                        if 'exec' in res:
                            print(f"   Exec: {res['exec']:.1f}% ({res.get('successful_actions_count', res.get('executed_actions_count', 0))}/{res.get('total_generated_actions', 0)} actions succeeded)")
                        print()
                        total_gcr += res['gcr']
                        count_gcr += 1
                
                if count_gcr > 0:
                    avg_gcr = total_gcr / count_gcr
                    print(f"   Average GCR: {avg_gcr:.1f}%")
            
            # Baseline Summary
            if baseline_results:
                print("\n📊 BASELINE RESULTS:")
                print("-" * 70)
                total_gcr = 0
                count_gcr = 0
                for res in baseline_results:
                    if 'gcr' in res:
                        status_icon = "✅" if res['status'] == "PASSED" else "❌"
                        print(f"{status_icon} {res['task']}")
                        print(f"   Status: {res['status']}")
                        print(f"   GCR: {res['gcr']:.1f}%")
                        if 'exec' in res:
                            print(f"   Exec: {res['exec']:.1f}% ({res.get('successful_actions_count', res.get('executed_actions_count', 0))}/{res.get('total_generated_actions', 0)} actions succeeded)")
                        print()
                        total_gcr += res['gcr']
                        count_gcr += 1
                
                if count_gcr > 0:
                    avg_gcr = total_gcr / count_gcr
                    print(f"   Average GCR: {avg_gcr:.1f}%")
            
            # Side-by-side Comparison
            if physical_guard_results and baseline_results:
                print("\n📊 SIDE-BY-SIDE COMPARISON:")
                print("=" * 70)
                print(f"{'Task':<40} {'Physical Guard':<20} {'Baseline':<20}")
                print("-" * 70)
                
                pg_dict = {r['task']: r for r in physical_guard_results}
                bl_dict = {r['task']: r for r in baseline_results}
                all_tasks = set(pg_dict.keys()) | set(bl_dict.keys())
                
                for task in sorted(all_tasks):
                    pg_res = pg_dict.get(task, {})
                    bl_res = bl_dict.get(task, {})
                    
                    pg_gcr = pg_res.get('gcr', 0) if 'gcr' in pg_res else None
                    bl_gcr = bl_res.get('gcr', 0) if 'gcr' in bl_res else None
                    
                    pg_str = f"{pg_gcr:.1f}%" if pg_gcr is not None else "N/A"
                    bl_str = f"{bl_gcr:.1f}%" if bl_gcr is not None else "N/A"
                    
                    if pg_gcr is not None and bl_gcr is not None:
                        if pg_gcr > bl_gcr:
                            pg_str = f"🏆 {pg_str}"
                        elif bl_gcr > pg_gcr:
                            bl_str = f"🏆 {bl_str}"
                    
                    print(f"{task[:38]:<40} {pg_str:<20} {bl_str:<20}")
                
                print("=" * 70)
            
            # Excel 파일로 저장
            project_root = Path(current_dir).parent
            output_file = project_root / f"FP{fp_number}_result.xlsx"
            
            print(f"\n💾 Excel 파일 저장 중...")
            save_results_to_excel(
                physical_guard_results,
                baseline_results,
                fp_number,
                str(output_file)
            )
            
        except Exception as e:
            print(f"❌ FloorPlan{fp_number} 평가 실패: {e}")
            import traceback
            traceback.print_exc()
            continue
    
    print(f"\n{'='*70}")
    print(f"✅ 모든 평가 완료! ({len(fp_folders)}개 FP)")
    print(f"{'='*70}")
    
    # Auto-select test_file based on FP number if not specified
    if not args.test_file or args.test_file == "data/final_test/FloorPlan1.json":
        test_file_path = os.path.join("data/final_test", f"FloorPlan{fp_number}.json")
        if os.path.exists(test_file_path):
            args.test_file = test_file_path
            print(f"📋 FP{fp_number}에 맞는 test_file 자동 선택: {args.test_file}")
        else:
            print(f"⚠️  Warning: {test_file_path} not found, using default: {args.test_file}")
    else:
        print(f"📋 Using specified test file: {args.test_file}")
    
    # Find JSON files in the specified folder/FP{num} directory
    folder_name = args.folder
    fp_folder = f"FP{fp_number}"
    print(f"📂 Searching in results/{folder_name}/{fp_folder} folder...")
    
    # Try absolute path first
    folder_physical_guard_pattern = os.path.join(current_dir, "../results", folder_name, fp_folder, "physical_guard_result_*.json")
    folder_baseline_pattern = os.path.join(current_dir, "../results", folder_name, fp_folder, "baseline_result_*.json")
    
    physical_guard_files = glob.glob(folder_physical_guard_pattern)
    baseline_files = glob.glob(folder_baseline_pattern)
    
    # Try relative path if absolute path didn't work
    if not physical_guard_files:
        folder_physical_guard_pattern = os.path.join("results", folder_name, fp_folder, "physical_guard_result_*.json")
        physical_guard_files = glob.glob(folder_physical_guard_pattern)
    
    if not baseline_files:
        folder_baseline_pattern = os.path.join("results", folder_name, fp_folder, "baseline_result_*.json")
        baseline_files = glob.glob(folder_baseline_pattern)
    
    # Fallback: Check parent folder (results/{folder_name}/) if FP{num} folder has no files
    if not physical_guard_files:
        search_pattern = os.path.join(current_dir, "../results", folder_name, "physical_guard_result_*.json")
        physical_guard_files = glob.glob(search_pattern)
        if not physical_guard_files:
            search_pattern = os.path.join("results", folder_name, "physical_guard_result_*.json")
            physical_guard_files = glob.glob(search_pattern)
    
    if not baseline_files:
        search_pattern = os.path.join(current_dir, "../results", folder_name, "baseline_result_*.json")
        baseline_files = glob.glob(search_pattern)
        if not baseline_files:
            search_pattern = os.path.join("results", folder_name, "baseline_result_*.json")
            baseline_files = glob.glob(search_pattern)
    
    # Final fallback: Check results folder directly
    if not physical_guard_files:
        search_pattern = os.path.join(current_dir, "../results", "physical_guard_result_*.json")
        physical_guard_files = glob.glob(search_pattern)
        if not physical_guard_files:
            search_pattern = "results/physical_guard_result_*.json"
            physical_guard_files = glob.glob(search_pattern)
    
    if not baseline_files:
        search_pattern = os.path.join(current_dir, "../results", "baseline_result_*.json")
        baseline_files = glob.glob(search_pattern)
        if not baseline_files:
            search_pattern = "results/baseline_result_*.json"
            baseline_files = glob.glob(search_pattern)
    
    # Use command line arguments if provided
    if args.physical_guard_json:
        physical_guard_files = [args.physical_guard_json]
    if args.baseline_json:
        baseline_files = [args.baseline_json]
    
    # Get latest files if multiple found
    physical_guard_file = None
    baseline_file = None
    
    if physical_guard_files:
        physical_guard_file = max(physical_guard_files, key=os.path.getmtime)
        print(f"📁 Physical Guard JSON: {physical_guard_file}")
    
    if baseline_files:
        baseline_file = max(baseline_files, key=os.path.getmtime)
        print(f"📁 Baseline JSON: {baseline_file}")
    
    if not physical_guard_file and not baseline_file:
        print("❌ No JSON files found.")
        print(f"   Searched for:")
        print(f"     - results/{folder_name}/{fp_folder}/physical_guard_result_*.json")
        print(f"     - results/{folder_name}/{fp_folder}/baseline_result_*.json")
        print(f"     - results/{folder_name}/physical_guard_result_*.json")
        print(f"     - results/{folder_name}/baseline_result_*.json")
        print(f"     - results/physical_guard_result_*.json")
        print(f"     - results/baseline_result_*.json")
        print(f"\n   Available folders in results/{folder_name}/:")
        folder_dir = os.path.join(current_dir, "../results", folder_name)
        if os.path.exists(folder_dir):
            for item in os.listdir(folder_dir):
                item_path = os.path.join(folder_dir, item)
                if os.path.isdir(item_path):
                    print(f"     - {item}")
        return
    
    # Load plans
    physical_guard_plans = {}
    baseline_plans = {}
    
    if physical_guard_file:
        print(f"\n📖 Loading Physical Guard plans from: {physical_guard_file}")
        physical_guard_plans = parse_json_plan_file(physical_guard_file)
        print(f"   Loaded {len(physical_guard_plans)} tasks")
    
    if baseline_file:
        print(f"\n📖 Loading Baseline plans from: {baseline_file}")
        baseline_plans = parse_json_plan_file(baseline_file)
        print(f"   Loaded {len(baseline_plans)} tasks")
    
    print("\n📋 Loading expected results...")
    expected_results = load_expected_results(args.test_file)
    print(f"   Expected tasks: {len(expected_results)}")
    for task_def in expected_results:
        print(f"     - {task_def['task']}")
    
    # Initialize Executors
    print("\n🤖 Initializing AI2-THOR executors...")
    physical_guard_exe = None
    baseline_exe = None
    
    # Determine scene name from FP number
    scene_name = f"FloorPlan{fp_number}"
    
    if physical_guard_plans:
        physical_guard_exe = AI2ThorExecutor(
            scene=scene_name,
            headless=False, 
            save_video=False 
        )
        physical_guard_exe.initialize()
        print(f"   ✓ Physical Guard executor initialized (Scene: {scene_name})")
    
    if baseline_plans:
        baseline_exe = AI2ThorExecutor(
            scene=scene_name,
            headless=False, 
            save_video=False 
        )
        baseline_exe.initialize()
        print(f"   ✓ Baseline executor initialized (Scene: {scene_name})")
    
    # Evaluate tasks
    physical_guard_results = []
    baseline_results = []
    
    for task_def in expected_results:
        task_name = task_def['task']
        task_key = task_name.strip().lower()
        
        print(f"\n{'='*70}")
        print(f"TASK: {task_name}")
        print(f"{'='*70}")
        
        # 공백/대소문자 무시 정규화 (sinkbasin vs sink basin 등)
        def _norm(s: str) -> str:
            return (s or "").lower().strip().replace(" ", "")

        # Evaluate Physical Guard
        if physical_guard_plans and physical_guard_exe:
            matched_key = None
            if task_key in physical_guard_plans:
                matched_key = task_key
            else:
                # 정규화 후 일치 시 매칭
                task_norm = _norm(task_name)
                for plan_key in physical_guard_plans.keys():
                    if _norm(plan_key) == task_norm:
                        matched_key = plan_key
                        print(f"  ℹ️ Using normalized match: '{plan_key}' for Physical Guard")
                        break
                if not matched_key:
                    # 기존 fuzzy (포함 관계)
                    for plan_key in physical_guard_plans.keys():
                        if task_key in plan_key or plan_key in task_key:
                            matched_key = plan_key
                            print(f"  ℹ️ Using fuzzy match: '{plan_key}' for Physical Guard")
                            break
            
            if matched_key:
                program_code = physical_guard_plans[matched_key]
                result = execute_and_evaluate_task(
                    physical_guard_exe, task_def, program_code, "Physical Guard"
                )
                physical_guard_results.append(result)
            else:
                print(f"  ⚠️ Physical Guard: Plan not found for '{task_name}'")
                if physical_guard_plans:
                    all_keys = list(physical_guard_plans.keys())
                    print(f"     (로드된 plan 키 {len(all_keys)}개: {all_keys})")
                physical_guard_results.append({
                    "task": task_name,
                    "method": "Physical Guard",
                    "status": "SKIPPED",
                    "reason": "No Plan"
                })
        
        # Evaluate Baseline
        if baseline_plans and baseline_exe:
            matched_key = None
            if task_key in baseline_plans:
                matched_key = task_key
            else:
                task_norm = _norm(task_name)
                for plan_key in baseline_plans.keys():
                    if _norm(plan_key) == task_norm:
                        matched_key = plan_key
                        print(f"  ℹ️ Using normalized match: '{plan_key}' for Baseline")
                        break
                if not matched_key:
                    for plan_key in baseline_plans.keys():
                        if task_key in plan_key or plan_key in task_key:
                            matched_key = plan_key
                            print(f"  ℹ️ Using fuzzy match: '{plan_key}' for Baseline")
                            break
            
            if matched_key:
                program_code = baseline_plans[matched_key]
                result = execute_and_evaluate_task(
                    baseline_exe, task_def, program_code, "Baseline"
                )
                baseline_results.append(result)
            else:
                print(f"  ⚠️ Baseline: Plan not found for '{task_name}'")
                if baseline_plans:
                    all_keys = list(baseline_plans.keys())
                    print(f"     (로드된 plan 키 {len(all_keys)}개: {all_keys})")
                baseline_results.append({
                    "task": task_name,
                    "method": "Baseline",
                    "status": "SKIPPED",
                    "reason": "No Plan"
                })
    
    # Print Comparison Summary
    print("\n" + "="*70)
    print("EVALUATION SUMMARY - COMPARISON")
    print("="*70)
    
    # Physical Guard Summary
    if physical_guard_results:
        print("\n📊 PHYSICAL GUARD RESULTS:")
        print("-" * 70)
        total_gcr = 0
        count_gcr = 0
        total_passed_conds = 0
        total_conds = 0
        
        for res in physical_guard_results:
            if 'gcr' in res:
                status_icon = "✅" if res['status'] == "PASSED" else "❌"
                print(f"{status_icon} {res['task']}")
                print(f"   Status: {res['status']}")
                print(f"   GCR: {res['gcr']:.1f}%")
                if 'exec' in res:
                    print(f"   Exec: {res['exec']:.1f}% ({res.get('successful_actions_count', res.get('executed_actions_count', 0))}/{res.get('total_generated_actions', 0)} actions succeeded)")
                if 'passed_conditions' in res and 'total_conditions' in res:
                    print(f"   Conditions: {res.get('passed_conditions', 0)}/{res.get('total_conditions', 0)} passed")
                print()
                total_gcr += res['gcr']
                count_gcr += 1
                total_passed_conds += res.get('passed_conditions', 0)
                total_conds += res.get('total_conditions', 0)
            else:
                print(f"⚠️  {res['task']}: {res.get('status', 'SKIPPED')}")
                print()
        
        if count_gcr > 0:
            avg_gcr = total_gcr / count_gcr
            print(f"   Average GCR: {avg_gcr:.1f}%")
            print(f"   Overall Conditions: {total_passed_conds}/{total_conds} passed")
    
    # Baseline Summary
    if baseline_results:
        print("\n📊 BASELINE RESULTS:")
        print("-" * 70)
        total_gcr = 0
        count_gcr = 0
        total_passed_conds = 0
        total_conds = 0
        
        for res in baseline_results:
            if 'gcr' in res:
                status_icon = "✅" if res['status'] == "PASSED" else "❌"
                print(f"{status_icon} {res['task']}")
                print(f"   Status: {res['status']}")
                print(f"   GCR: {res['gcr']:.1f}%")
                if 'exec' in res:
                    print(f"   Exec: {res['exec']:.1f}% ({res.get('successful_actions_count', res.get('executed_actions_count', 0))}/{res.get('total_generated_actions', 0)} actions succeeded)")
                if 'passed_conditions' in res and 'total_conditions' in res:
                    print(f"   Conditions: {res.get('passed_conditions', 0)}/{res.get('total_conditions', 0)} passed")
                print()
                total_gcr += res['gcr']
                count_gcr += 1
                total_passed_conds += res.get('passed_conditions', 0)
                total_conds += res.get('total_conditions', 0)
            else:
                print(f"⚠️  {res['task']}: {res.get('status', 'SKIPPED')}")
                print()
        
        if count_gcr > 0:
            avg_gcr = total_gcr / count_gcr
            print(f"   Average GCR: {avg_gcr:.1f}%")
            print(f"   Overall Conditions: {total_passed_conds}/{total_conds} passed")
    
    # Side-by-side Comparison
    if physical_guard_results and baseline_results:
        print("\n📊 SIDE-BY-SIDE COMPARISON:")
        print("=" * 70)
        print(f"{'Task':<40} {'Physical Guard':<20} {'Baseline':<20}")
        print("-" * 70)
        
        # Create task mapping
        pg_dict = {r['task']: r for r in physical_guard_results}
        bl_dict = {r['task']: r for r in baseline_results}
        
        all_tasks = set(pg_dict.keys()) | set(bl_dict.keys())
        
        for task in sorted(all_tasks):
            pg_res = pg_dict.get(task, {})
            bl_res = bl_dict.get(task, {})
            
            pg_gcr = pg_res.get('gcr', 0) if 'gcr' in pg_res else None
            bl_gcr = bl_res.get('gcr', 0) if 'gcr' in bl_res else None
            
            pg_str = f"{pg_gcr:.1f}%" if pg_gcr is not None else "N/A"
            bl_str = f"{bl_gcr:.1f}%" if bl_gcr is not None else "N/A"
            
            # Highlight winner
            if pg_gcr is not None and bl_gcr is not None:
                if pg_gcr > bl_gcr:
                    pg_str = f"🏆 {pg_str}"
                elif bl_gcr > pg_gcr:
                    bl_str = f"🏆 {bl_str}"
            
            print(f"{task[:38]:<40} {pg_str:<20} {bl_str:<20}")
        
        print("=" * 70)

if __name__ == "__main__":
    main()
