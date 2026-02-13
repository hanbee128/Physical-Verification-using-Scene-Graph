#!/usr/bin/env python3
# Ablation Study: 한 번의 실행에서 3가지 방법 비교
# (1) Baseline: Baseline(ProgPrompt).py → plan 생성
# (2) Ablation: 논리적 검증(generate_program) + 물리적 검증(Physical Guard + Recovery) → plan 생성
# (3) Physical Guard: physical_guard.py 전체 실행 → plan 생성
# 이후 3가지 plan을 동일 조건에서 실행·평가(GCR/Exec)하여 비교

import os
import sys
import json
import glob
import shutil
import subprocess
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
from collections import defaultdict

# scripts 디렉터리를 path에 추가 (physical_guard, evaluate_results import)
_script_dir = Path(__file__).parent
_project_root = _script_dir.parent
# batch_evaluation과 동일 구조: results_ablation/Kitchen/FP1, LivingRoom/FP216 등
ABLATION_RESULTS_BASE = "results_ablation"
if str(_script_dir) not in sys.path:
    sys.path.insert(0, str(_script_dir))

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter


def get_folder_from_fp_number(fp_number: int) -> str:
    if fp_number in [1, 2]:
        return "Kitchen"
    elif fp_number in [216, 224]:
        return "LivingRoom"
    elif fp_number in [325, 326]:
        return "BedRoom"
    elif fp_number in [403, 425]:
        return "BathRoom"
    return "Kitchen"


def load_tasks_from_fp(project_root: Path, fp_number: int) -> List[Dict[str, Any]]:
    path = project_root / "data" / "final_test" / f"FloorPlan{fp_number}.json"
    if not path.exists():
        return []
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, list):
        return data
    if isinstance(data, dict) and "task" in data:
        return [data]
    return []


def run_baseline_subprocess(fp_number: int, task_file: str, folder: str) -> bool:
    """Baseline(ProgPrompt).py를 subprocess로 실행. batch_evaluation과 동일."""
    baseline_script = _script_dir / "Baseline(ProgPrompt).py"
    if not baseline_script.exists():
        print(f"⚠️  Baseline 스크립트를 찾을 수 없습니다: {baseline_script}")
        return False
    out_dir = _project_root / ABLATION_RESULTS_BASE / folder / f"FP{fp_number}"
    out_dir.mkdir(parents=True, exist_ok=True)
    cmd = [
        sys.executable,
        str(baseline_script),
        "--scene-number", str(fp_number),
        "--output-dir", str(out_dir),
        "--task-file", task_file,
    ]
    try:
        process = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            universal_newlines=True,
            bufsize=1,
            cwd=_project_root,
        )
        for line in process.stdout:
            print(line.rstrip())
        process.wait()
        return process.returncode == 0
    except Exception as e:
        print(f"❌ Baseline 실행 실패: {e}")
        return False


def find_latest_baseline_json(fp_number: int, folder: str) -> str:
    """results_ablation/{folder}/FP{fp_number}/ 내 최신 baseline_result_*.json 경로 반환."""
    base = _project_root / ABLATION_RESULTS_BASE / folder / f"FP{fp_number}"
    if not base.exists():
        return ""
    files = glob.glob(str(base / "baseline_result_*.json"))
    if not files:
        return ""
    return max(files, key=os.path.getmtime)


def find_latest_physical_guard_json(fp_number: int, folder: str) -> str:
    """results_ablation/{folder}/FP{fp_number}/ 내 최신 physical_guard_result_*.json 경로 반환."""
    base = _project_root / ABLATION_RESULTS_BASE / folder / f"FP{fp_number}"
    if not base.exists():
        return ""
    files = glob.glob(str(base / "physical_guard_result_*.json"))
    if not files:
        return ""
    return max(files, key=os.path.getmtime)


def run_physical_guard_subprocess(fp_number: int, task_file: str, folder: str) -> bool:
    """physical_guard.py를 subprocess로 실행. 결과를 results_ablation/{folder}/FP{fp_number}/ 에 저장."""
    physical_guard_script = _script_dir / "physical_guard.py"
    if not physical_guard_script.exists():
        print(f"⚠️  physical_guard.py를 찾을 수 없습니다: {physical_guard_script}")
        return False
    out_dir = _project_root / ABLATION_RESULTS_BASE / folder / f"FP{fp_number}"
    out_dir.mkdir(parents=True, exist_ok=True)
    cmd = [
        sys.executable,
        str(physical_guard_script),
        "--scene-number", str(fp_number),
        "--output-dir", str(out_dir),
        "--task-file", task_file,
    ]
    try:
        process = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            universal_newlines=True,
            bufsize=1,
            cwd=_project_root,
        )
        for line in process.stdout:
            print(line.rstrip())
        process.wait()
        return process.returncode == 0
    except Exception as e:
        print(f"❌ physical_guard.py 실행 실패: {e}")
        return False


def run_ablation_baseline_pg_recovery(
    fp_number: int,
    folder: str,
    test_file: str,
) -> str:
    """
    physical_guard와 동일: 논리적 검증(generate_program) → 물리적 검증(Physical Guard + Recovery).
    누락 task LLM 감지/추가 생성만 하지 않음. 결과를 ablation_baseline_pg_recovery_result.json에 저장.
    """
    from evaluate_results import load_expected_results
    from physical_guard import (
        load_scene_graph,
        generate_final_plan_with_physical_verification,
        generate_program,
        build_prompt,
        parse_info_txt,
        DEFAULT_EXAMPLES,
        AI2THOR_ACTIONS,
        DEFAULT_FLOORPLAN1_OBJECTS,
    )
    from openai import OpenAI

    expected_results = load_expected_results(test_file)
    scene_graph_path = _script_dir / f"scene_graph_structured_FloorPlan{fp_number}.json"
    if not scene_graph_path.exists():
        print(f"⚠️  Scene Graph 없음: {scene_graph_path}")
        return ""

    # physical_guard와 동일: info.txt에서 액션/객체 로드 (FP{num}_info.txt 우선)
    info_path = _project_root / "data" / "all_plans_env0" / f"FP{fp_number}_info.txt"
    if not info_path.exists():
        info_path = _project_root / "data" / "all_plans_env0" / "info.txt"
    if info_path.exists():
        actions, objects = parse_info_txt(str(info_path))
        if not actions:
            actions = AI2THOR_ACTIONS
        if not objects:
            objects = sorted(DEFAULT_FLOORPLAN1_OBJECTS)
        else:
            objects = sorted(objects)
    else:
        actions = AI2THOR_ACTIONS
        objects = sorted(DEFAULT_FLOORPLAN1_OBJECTS)
    examples = dict(list(DEFAULT_EXAMPLES.items()))
    prompt = build_prompt(objects, actions, examples, max_examples=3)
    ollama_url = "http://localhost:11434/v1"
    model = "llama3"
    temperature = 0.0
    max_tokens = 700

    out_dir = _project_root / ABLATION_RESULTS_BASE / folder / f"FP{fp_number}"
    out_dir.mkdir(parents=True, exist_ok=True)
    updated_sg_path = out_dir / "updated_scene_graph_ablation.json"
    shutil.copy2(str(scene_graph_path), str(updated_sg_path))
    scene_graph = load_scene_graph(str(updated_sg_path))

    controller = None
    try:
        from ai2thor.controller import Controller
        scene_name = f"FloorPlan{fp_number}_physics"
        controller = Controller(
            agentMode="arm",
            scene=scene_name,
            gridSize=0.25,
            snapToGrid=False,
            rotateStepDegrees=90,
            visibilityDistance=1.5,
            renderInstanceSegmentation=False,
            renderDepthImage=False,
            renderSemanticSegmentation=False,
            width=300,
            height=300,
            fieldOfView=90,
        )
    except Exception as e:
        print(f"  ⚠️  Controller 초기화 실패 (NavMesh 제한): {e}")

    ablation_plans = {}
    for task_def in expected_results:
        task_name = task_def.get("task", "")
        if not task_name:
            continue

        # task마다 scene graph 초기 상태로 리셋 (독립 실행)
        shutil.copy2(str(scene_graph_path), str(updated_sg_path))
        scene_graph = load_scene_graph(str(updated_sg_path))

        try:
            # physical_guard와 동일: 논리적 검증(LLM) → 물리적 검증 (누락 task LLM 단계만 생략)
            client = OpenAI(base_url=ollama_url, api_key="ollama")
            initial_program = generate_program(
                client=client,
                model=model,
                base_prompt=prompt,
                task=task_name,
                temperature=temperature,
                max_tokens=max_tokens,
            )
            final_program, _ = generate_final_plan_with_physical_verification(
                task=task_name,
                initial_program=initial_program,
                scene_graph=scene_graph,
                controller=controller,
                client=client,
                model=model,
                scene_graph_path=str(updated_sg_path),
            )
            if final_program:
                ablation_plans[task_name] = final_program
        except Exception as e:
            print(f"  ⚠️  Physical verification 실패 '{task_name}': {e}")

    if controller is not None:
        try:
            controller.stop()
        except Exception:
            pass

    out_path = out_dir / "ablation_baseline_pg_recovery_result.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(ablation_plans, f, indent=2, ensure_ascii=False)

    # Baseline과 동일 형식으로 txt 저장 (Task: ... + 구분선 + 프로그램 코드)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    txt_path = out_dir / f"ablation_baseline_pg_recovery_result_{timestamp}.txt"
    with open(txt_path, "w", encoding="utf-8") as f:
        for task, program in ablation_plans.items():
            f.write(f"Task: {task}\n")
            f.write("=" * 80 + "\n")
            f.write(program)
            f.write("\n\n" + "=" * 80 + "\n\n")
    print(f"✅ Ablation(BL+PG+Rec) plan txt 저장: {txt_path}")

    return str(out_path)


def _match_task_to_plan(task_key: str, task_name: str, plans: Dict[str, Any], _norm) -> Optional[str]:
    """task에 대응하는 plan 키 반환."""
    if task_key in plans:
        return task_key
    task_norm = _norm(task_name)
    for plan_key in plans.keys():
        if _norm(plan_key) == task_norm:
            return plan_key
    for plan_key in plans.keys():
        if task_key in plan_key or plan_key in task_key:
            return plan_key
    return None


def run_evaluate_three_methods(
    fp_number: int,
    folder: str,
    test_file: str,
    baseline_json_path: str,
    ablation_json_path: str,
    physical_guard_json_path: str,
) -> Dict[str, Any]:
    """Baseline / Ablation(BL+PG+Rec) / Physical Guard 3가지 plan을 실행·평가하여 비교."""
    from evaluate_results import load_expected_results, parse_json_plan_file, execute_and_evaluate_task
    from ai2thor_connector_ithor import AI2ThorExecutor

    baseline_plans = parse_json_plan_file(baseline_json_path)
    ablation_plans = parse_json_plan_file(ablation_json_path)
    pg_plans = parse_json_plan_file(physical_guard_json_path)
    expected_results = load_expected_results(test_file)
    scene_name = f"FloorPlan{fp_number}"

    def _norm(s: str) -> str:
        return (s or "").lower().strip().replace(" ", "")

    baseline_exe = None
    ablation_exe = None
    pg_exe = None
    if baseline_plans:
        baseline_exe = AI2ThorExecutor(scene=scene_name, headless=False, save_video=False)
        baseline_exe.initialize()
    if ablation_plans:
        ablation_exe = AI2ThorExecutor(scene=scene_name, headless=False, save_video=False)
        ablation_exe.initialize()
    if pg_plans:
        pg_exe = AI2ThorExecutor(scene=scene_name, headless=False, save_video=False)
        pg_exe.initialize()

    baseline_results = []
    ablation_results = []
    physical_guard_results = []
    for task_def in expected_results:
        task_name = task_def["task"]
        task_key = task_name.strip().lower()

        if baseline_plans and baseline_exe:
            matched = _match_task_to_plan(task_key, task_name, baseline_plans, _norm)
            if matched:
                baseline_results.append(
                    execute_and_evaluate_task(
                        baseline_exe, task_def, baseline_plans[matched], "Baseline"
                    )
                )
            else:
                baseline_results.append({
                    "task": task_name, "method": "Baseline", "status": "SKIPPED", "reason": "No Plan"
                })

        if ablation_plans and ablation_exe:
            matched = _match_task_to_plan(task_key, task_name, ablation_plans, _norm)
            if matched:
                ablation_results.append(
                    execute_and_evaluate_task(
                        ablation_exe, task_def, ablation_plans[matched], "Ablation"
                    )
                )
            else:
                ablation_results.append({
                    "task": task_name, "method": "Ablation", "status": "SKIPPED", "reason": "No Plan"
                })

        if pg_plans and pg_exe:
            matched = _match_task_to_plan(task_key, task_name, pg_plans, _norm)
            if matched:
                physical_guard_results.append(
                    execute_and_evaluate_task(
                        pg_exe, task_def, pg_plans[matched], "Physical Guard"
                    )
                )
            else:
                physical_guard_results.append({
                    "task": task_name, "method": "Physical Guard", "status": "SKIPPED", "reason": "No Plan"
                })

    for exe in (baseline_exe, ablation_exe, pg_exe):
        if exe:
            try:
                exe.close()
            except Exception:
                pass

    result_dict = {
        "baseline": baseline_results,
        "ablation": ablation_results,
        "physical_guard": physical_guard_results,
    }
    output_json = _project_root / ABLATION_RESULTS_BASE / folder / f"FP{fp_number}" / "_ablation_eval_result.json"
    output_json.parent.mkdir(parents=True, exist_ok=True)
    with open(output_json, "w", encoding="utf-8") as f:
        json.dump(result_dict, f, indent=2, ensure_ascii=False)
    return result_dict


def calculate_averages(all_results: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Baseline / Ablation / Physical Guard 3가지 평균 계산."""
    method_keys = ("baseline", "ablation", "physical_guard")
    task_results = {k: defaultdict(list) for k in method_keys}
    task_exec = {k: defaultdict(list) for k in method_keys}
    for run_result in all_results:
        for method in method_keys:
            for r in run_result.get(method, []):
                task_name = r.get("task", "Unknown")
                if "gcr" in r:
                    task_results[method][task_name].append(r["gcr"])
                if "exec" in r:
                    task_exec[method][task_name].append(r["exec"])
    avg_by_method = {}
    for method in method_keys:
        avg_by_method[method] = {}
        for task_name, gcr_list in task_results[method].items():
            exec_list = task_exec[method].get(task_name, [])
            avg_by_method[method][task_name] = {
                "avg_gcr": sum(gcr_list) / len(gcr_list),
                "gcr_list": gcr_list,
                "avg_exec": sum(exec_list) / len(exec_list) if exec_list else 0.0,
                "exec_list": exec_list,
                "count": len(gcr_list),
            }
    scene_avg = {}
    for method in method_keys:
        gcr_vals = [r["avg_gcr"] for r in avg_by_method[method].values()]
        exec_vals = [r["avg_exec"] for r in avg_by_method[method].values()]
        scene_avg[method] = sum(gcr_vals) / len(gcr_vals) if gcr_vals else 0.0
        scene_avg[f"{method}_exec"] = sum(exec_vals) / len(exec_vals) if exec_vals else 0.0
    return {**avg_by_method, "scene_avg": scene_avg}


def save_to_excel(
    all_scene_results: Dict[str, Dict[str, Any]],
    all_run_results: Dict[str, List[Dict[str, Any]]],
    output_file: str,
    fp_number: Optional[int] = None,
    task_index: Optional[int] = None,
) -> None:
    """Baseline / Ablation / Physical Guard 3가지 비교 Excel 저장."""
    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    method_keys = ("baseline", "ablation", "physical_guard")
    method_labels = ("Baseline", "Ablation", "Physical Guard")

    def _sheet_title_run(fp_name: str, run_num: int) -> str:
        if fp_number is not None and task_index is not None:
            return f"FP{fp_number}_Task{task_index}_R{run_num}"[:31]
        return f"{fp_name}_R{run_num}"[:31]

    def _sheet_title_avg(fp_name: str) -> str:
        if fp_number is not None and task_index is not None:
            return f"FP{fp_number}_Task{task_index}_Avg"[:31]
        return f"{fp_name}_Avg"[:31]

    # Run별 시트: Task | Baseline GCR | Ablation GCR | PG GCR | Baseline Exec | Ablation Exec | PG Exec
    for fp_name, run_list in all_run_results.items():
        for run_data in run_list:
            run_num = run_data["run_number"]
            result = run_data["result"]
            sheet_name = _sheet_title_run(fp_name, run_num)
            ws = wb.create_sheet(title=sheet_name)
            headers = ["Task"] + [f"{lbl} GCR (%)" for lbl in method_labels] + [f"{lbl} Exec (%)" for lbl in method_labels]
            for col_idx, header in enumerate(headers, 1):
                c = ws.cell(row=1, column=col_idx, value=header)
                c.fill = header_fill
                c.font = header_font
                c.alignment = center_align
            by_task = {k: {r["task"]: r for r in result.get(k, [])} for k in method_keys}
            all_tasks = set()
            for d in by_task.values():
                all_tasks |= set(d.keys())
            row = 2
            for task_name in sorted(all_tasks):
                ws.cell(row=row, column=1, value=task_name)
                for col, k in enumerate(method_keys, 2):
                    res = by_task[k].get(task_name, {})
                    ws.cell(row=row, column=col, value=round(res.get("gcr", 0.0), 2))
                for col, k in enumerate(method_keys, 2 + len(method_keys)):
                    res = by_task[k].get(task_name, {})
                    ws.cell(row=row, column=col, value=round(res.get("exec", 0.0), 2))
                row += 1
            ws.column_dimensions["A"].width = 40
            for c in range(2, 8):
                ws.column_dimensions[get_column_letter(c)].width = 16

    # Scene별 평균 시트
    for fp_name, scene_data in all_scene_results.items():
        ws = wb.create_sheet(title=_sheet_title_avg(fp_name))
        headers = ["Task"] + [f"{lbl} 평균 GCR (%)" for lbl in method_labels] + [f"{lbl} 평균 Exec (%)" for lbl in method_labels]
        for col_idx, header in enumerate(headers, 1):
            c = ws.cell(row=1, column=col_idx, value=header)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center_align
        scene_avg = scene_data.get("scene_avg", {})
        all_tasks = set()
        for k in method_keys:
            all_tasks |= set(scene_data.get(k, {}).keys())
        row = 2
        for task_name in sorted(all_tasks):
            ws.cell(row=row, column=1, value=task_name)
            for col, k in enumerate(method_keys, 2):
                info = scene_data.get(k, {}).get(task_name, {})
                ws.cell(row=row, column=col, value=round(info.get("avg_gcr", 0.0), 2))
            for col, k in enumerate(method_keys, 2 + len(method_keys)):
                info = scene_data.get(k, {}).get(task_name, {})
                ws.cell(row=row, column=col, value=round(info.get("avg_exec", 0.0), 2))
            row += 1
        row += 1
        ws.cell(row=row, column=1, value="Scene 평균")
        for col, k in enumerate(method_keys, 2):
            ws.cell(row=row, column=col, value=round(scene_avg.get(k, 0), 2))
        for col, k in enumerate(method_keys, 2 + len(method_keys)):
            ws.cell(row=row, column=col, value=round(scene_avg.get(f"{k}_exec", 0), 2))
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws.column_dimensions["A"].width = 40
        for c in range(2, 8):
            ws.column_dimensions[get_column_letter(c)].width = 18

    # Summary 시트: 3방법 GCR / Exec 비교
    summary_ws = wb.create_sheet(title="Summary", index=0)
    summary_headers = ["Scene"] + [f"{lbl} GCR (%)" for lbl in method_labels] + [f"{lbl} Exec (%)" for lbl in method_labels]
    for col_idx, header in enumerate(summary_headers, 1):
        c = summary_ws.cell(row=1, column=col_idx, value=header)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center_align
    row = 2
    n = len(all_scene_results)
    for fp_name in sorted(all_scene_results.keys()):
        scene_avg = all_scene_results[fp_name]["scene_avg"]
        summary_ws.cell(row=row, column=1, value=fp_name)
        for col, k in enumerate(method_keys, 2):
            summary_ws.cell(row=row, column=col, value=round(scene_avg.get(k, 0), 2))
        for col, k in enumerate(method_keys, 2 + len(method_keys)):
            summary_ws.cell(row=row, column=col, value=round(scene_avg.get(f"{k}_exec", 0), 2))
        row += 1
    row += 1
    summary_ws.cell(row=row, column=1, value="전체 평균")
    if n > 0:
        for col, k in enumerate(method_keys, 2):
            summary_ws.cell(row=row, column=col, value=round(sum(r["scene_avg"].get(k, 0) for r in all_scene_results.values()) / n, 2))
        for col, k in enumerate(method_keys, 2 + len(method_keys)):
            summary_ws.cell(row=row, column=col, value=round(sum(r["scene_avg"].get(f"{k}_exec", 0) for r in all_scene_results.values()) / n, 2))
    else:
        for col in range(2, 8):
            summary_ws.cell(row=row, column=col, value=0)
    for col in range(1, 8):
        cell = summary_ws.cell(row=row, column=col)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    for col in range(1, 8):
        summary_ws.column_dimensions[get_column_letter(col)].width = 20
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    wb.save(output_file)
    print(f"\n✅ Excel 저장 완료: {output_file}")


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Ablation Study: Baseline / Ablation / Physical Guard 3방법 plan 생성·실행·평가 비교")
    parser.add_argument("--fp-number", type=int, default=None, help="FloorPlan 번호 (미지정 시 1, 2, 216, 224, 325, 326, 403, 425)")
    parser.add_argument("--num-runs", type=int, default=10, help="실행 횟수 (기본 10)")
    args = parser.parse_args()

    floor_plans = [args.fp_number] if args.fp_number else [1, 2, 216, 224, 325, 326, 403, 425]
    num_runs = args.num_runs

    # --fp-number 지정 시: 해당 FP의 Task 목록을 보여주고 번호로 선택 (0=전체, 1~N=단일 Task)
    selected_task_index = None
    single_task_path = None
    if args.fp_number:
        task_list = load_tasks_from_fp(_project_root, args.fp_number)
        if not task_list:
            print(f"❌ FloorPlan{args.fp_number}의 Task 파일을 읽을 수 없거나 비어 있습니다.")
            sys.exit(1)
        n_tasks = len(task_list)
        print(f"\n📋 FloorPlan{args.fp_number} 의 Task 목록 (총 {n_tasks}개):")
        for i, t in enumerate(task_list, 1):
            task_str = t.get("task", "(task 없음)")
            print(f"   {i}. {task_str}")
        while True:
            try:
                sel = input(f"\n실행할 Task 번호를 선택하세요 (0=해당 FP 전체 Task, 1-{n_tasks}=단일 Task): ").strip()
                task_index = int(sel)
                if task_index == 0:
                    selected_task_index = None
                    single_task_path = None
                    print(f"   선택: 0 — 해당 FP 전체 Task {n_tasks}개 {num_runs}번 실행")
                    break
                if 1 <= task_index <= n_tasks:
                    selected_task_def = task_list[task_index - 1]
                    single_task_path = _project_root / "data" / "final_test" / f"_single_task_fp{args.fp_number}_task{task_index}.json"
                    single_task_path.parent.mkdir(parents=True, exist_ok=True)
                    with open(single_task_path, "w", encoding="utf-8") as f:
                        json.dump([selected_task_def], f, indent=2, ensure_ascii=False)
                    selected_task_index = task_index
                    print(f"   선택: Task {task_index} — {num_runs}번 실행 후 결과 저장")
                    break
            except ValueError:
                pass
            print(f"   잘못된 입력입니다. 0 또는 1~{n_tasks} 사이 숫자를 입력하세요.")

    print("\n" + "=" * 70)
    print("🔬 Ablation Study: Baseline / Ablation(BL+PG+Rec) / Physical Guard 3방법 비교")
    print(f"   FloorPlans: {floor_plans}, Runs: {num_runs}")
    print("=" * 70)

    all_scene_results = {}
    all_run_results = defaultdict(list)
    start_time = datetime.now()

    try:
        for fp_number in floor_plans:
            folder = get_folder_from_fp_number(fp_number)
            # 단일 Task 선택 시 해당 task만 담긴 JSON 사용, 아니면 전체 FloorPlan JSON 사용
            if args.fp_number and selected_task_index is not None and single_task_path is not None:
                test_file = str(single_task_path)
            else:
                test_file = str(_project_root / "data" / "final_test" / f"FloorPlan{fp_number}.json")
            if not Path(test_file).exists():
                print(f"⚠️  테스트 파일 없음: {test_file}")
                continue
            fp_name = f"FloorPlan{fp_number}"
            all_run_results[fp_name] = []

            for run_num in range(1, num_runs + 1):
                print(f"\n{'='*70}")
                print(f"📌 Run {run_num}/{num_runs} - {fp_name}")
                print("=" * 70)

                # 1) Baseline 실행
                print("\n📝 Step 1: Baseline 실행 중...")
                if not run_baseline_subprocess(fp_number, test_file, folder):
                    print("⚠️  Baseline 실패, 해당 run 스킵")
                    continue
                baseline_json = find_latest_baseline_json(fp_number, folder)
                if not baseline_json:
                    print("⚠️  Baseline JSON을 찾을 수 없음")
                    continue

                # 2) Ablation: 논리적 검증(generate_program) + 물리적 검증(Physical Guard + Recovery)
                print("\n📝 Step 2: Ablation (논리+물리 검증) 적용 중...")
                ablation_json = run_ablation_baseline_pg_recovery(
                    fp_number=fp_number,
                    folder=folder,
                    test_file=test_file,
                )
                if not ablation_json:
                    print("⚠️  Ablation 결과 저장 실패")
                    continue

                # 3) Physical Guard 전체 스크립트 실행 (plan 생성)
                print("\n📝 Step 3: physical_guard.py 실행 중...")
                if not run_physical_guard_subprocess(fp_number, test_file, folder):
                    print("⚠️  physical_guard.py 실패, 해당 run 스킵")
                    continue
                pg_json = find_latest_physical_guard_json(fp_number, folder)
                if not pg_json:
                    print("⚠️  Physical Guard JSON을 찾을 수 없음")
                    continue

                # 4) 3가지 plan 실행 및 평가 (Baseline / Ablation / Physical Guard)
                print("\n📝 Step 4: 실행 및 평가 (3방법 비교)...")
                result = run_evaluate_three_methods(
                    fp_number=fp_number,
                    folder=folder,
                    test_file=test_file,
                    baseline_json_path=baseline_json,
                    ablation_json_path=ablation_json,
                    physical_guard_json_path=pg_json,
                )
                all_run_results[fp_name].append({"run_number": run_num, "result": result})
                print(f"✅ Run {run_num} 완료")

            if all_run_results[fp_name]:
                avg_results = calculate_averages([r["result"] for r in all_run_results[fp_name]])
                all_scene_results[fp_name] = avg_results
                sav = avg_results["scene_avg"]
                print(f"\n✅ {fp_name} 완료 (평균 GCR — Baseline: {sav['baseline']:.2f}%, Ablation: {sav['ablation']:.2f}%, Physical Guard: {sav['physical_guard']:.2f}%)")

        if all_scene_results:
            out_dir = _project_root / ABLATION_RESULTS_BASE
            out_dir.mkdir(parents=True, exist_ok=True)
            # batch_evaluation과 동일: 단일 task 선택 시 FP번호_task번호, 전체(0) 선택 시 FP번호만
            if args.fp_number and selected_task_index is not None:
                output_file = out_dir / f"FP{args.fp_number}_task{selected_task_index}_ablation_result.xlsx"
            elif args.fp_number:
                output_file = out_dir / f"FP{args.fp_number}_ablation_result.xlsx"
            else:
                output_file = out_dir / "ablation_baseline_pg_recovery_result.xlsx"
            save_to_excel(
                all_scene_results,
                dict(all_run_results),
                str(output_file),
                fp_number=args.fp_number if args.fp_number else None,
                task_index=selected_task_index,
            )
            print(f"   저장된 파일: {output_file}")

        elapsed = (datetime.now() - start_time).total_seconds()
        print("\n" + "=" * 70)
        print("✅ Ablation Study: Baseline / Ablation / Physical Guard 3방법 비교 완료")
        print(f"⏱️  소요 시간: {int(elapsed // 60)}분 {int(elapsed % 60)}초")
        print("=" * 70)

    finally:
        # 단일 Task용 임시 JSON 삭제
        if single_task_path is not None and Path(single_task_path).exists():
            try:
                Path(single_task_path).unlink()
            except Exception:
                pass


if __name__ == "__main__":
    main()
