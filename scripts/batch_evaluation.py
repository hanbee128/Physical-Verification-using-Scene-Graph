#!/usr/bin/env python3
"""
각 FloorPlan을 10번씩 실행하여 평균 GCR을 계산하고 Excel 파일로 저장하는 스크립트

실행 순서:
1. FloorPlan1.json, FloorPlan216.json, FloorPlan325.json, FloorPlan403.json 각각 10번 실행
2. 각 task마다 10번의 시도 평균 GCR 기록
3. 각 Scene마다 평균 GCR을 Baseline과 비교
4. result_half_test.xlsx 파일에 저장
"""

import os
import sys
import json
import subprocess
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any
from collections import defaultdict

# Excel 파일 생성을 위한 라이브러리
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl이 설치되지 않았습니다. 설치 중...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

# evaluate_results.py는 subprocess로 실행하므로 import 불필요

def get_folder_from_fp_number(fp_number: int) -> str:
    """FloorPlan 번호에 따라 폴더 이름을 반환"""
    if fp_number in [1, 2]:
        return "Kitchen"
    elif fp_number in [216, 224]:
        return "LivingRoom"
    elif fp_number in [325, 326]:
        return "BedRoom"
    elif fp_number in [403, 425]:
        return "BathRoom"
    else:
        return "Kitchen"


def load_tasks_from_fp(project_root: Path, fp_number: int) -> List[Dict[str, Any]]:
    """
    FloorPlan JSON에서 task 목록(전체 객체 리스트)을 로드합니다.
    Returns:
        [{"task": str, "object_states": ...}, ...] 또는 []
    """
    path = project_root / "data" / "final_test" / f"FloorPlan{fp_number}.json"
    if not path.exists():
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return data
        if isinstance(data, dict) and "task" in data:
            return [data]
        return []
    except Exception:
        return []

def find_result_files(fp_number: int, folder: str, after_time: datetime = None) -> tuple[str, str]:
    """
    Physical Guard와 Baseline 결과 JSON 파일을 찾습니다.
    
    Args:
        fp_number: FloorPlan 번호
        folder: 폴더 이름
        after_time: 이 시간 이후에 생성된 파일만 찾기 (None이면 가장 최신 파일)
    
    Returns:
        (physical_guard_file, baseline_file) 경로 튜플
    """
    import glob
    
    script_dir = Path(__file__).parent
    fp_folder = f"FP{fp_number}"
    
    # Physical Guard 파일 찾기
    patterns = [
        script_dir.parent / "results" / folder / fp_folder / "physical_guard_result_*.json",
        script_dir.parent / "results" / folder / "physical_guard_result_*.json",
        script_dir.parent / "results" / "physical_guard_result_*.json"
    ]
    
    physical_guard_file = None
    for pattern in patterns:
        files = glob.glob(str(pattern))
        if files:
            if after_time:
                # after_time 이후에 생성된 파일만 필터링
                filtered_files = [
                    f for f in files 
                    if datetime.fromtimestamp(os.path.getmtime(f)) >= after_time
                ]
                if filtered_files:
                    physical_guard_file = max(filtered_files, key=os.path.getmtime)
            else:
                physical_guard_file = max(files, key=os.path.getmtime)
            if physical_guard_file:
                break
    
    # Baseline 파일 찾기
    patterns = [
        script_dir.parent / "results" / folder / fp_folder / "baseline_result_*.json",
        script_dir.parent / "results" / folder / "baseline_result_*.json",
        script_dir.parent / "results" / "baseline_result_*.json"
    ]
    
    baseline_file = None
    for pattern in patterns:
        files = glob.glob(str(pattern))
        if files:
            if after_time:
                # after_time 이후에 생성된 파일만 필터링
                filtered_files = [
                    f for f in files 
                    if datetime.fromtimestamp(os.path.getmtime(f)) >= after_time
                ]
                if filtered_files:
                    baseline_file = max(filtered_files, key=os.path.getmtime)
            else:
                baseline_file = max(files, key=os.path.getmtime)
            if baseline_file:
                break
    
    return physical_guard_file, baseline_file

def run_physical_guard_subprocess(fp_number: int, task_file: str, folder: str = None) -> bool:
    """physical_guard.py를 subprocess로 실행 (실시간 출력).
    folder를 주면 결과를 results/{folder}/FP{fp_number}/ 에 저장합니다."""
    script_dir = Path(__file__).parent
    project_root = script_dir.parent
    physical_guard_script = script_dir / "physical_guard.py"
    
    cmd = [sys.executable, str(physical_guard_script), "--scene-number", str(fp_number)]
    if folder:
        out_dir = project_root / "results" / folder / f"FP{fp_number}"
        out_dir.mkdir(parents=True, exist_ok=True)
        cmd.extend(["--output-dir", str(out_dir)])
    if task_file:
        cmd.extend(["--task-file", task_file])
    
    try:
        process = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            universal_newlines=True,
            bufsize=1,
            cwd=script_dir.parent
        )
        
        # 실시간으로 출력
        for line in process.stdout:
            line = line.rstrip()
            print(line)
        
        process.wait()
        return process.returncode == 0
    except Exception as e:
        print(f"❌ physical_guard.py 실행 실패: {e}")
        return False

def run_evaluate_results_subprocess(fp_number: int, folder: str, test_file: str) -> Dict[str, Any]:
    """
    evaluate_results.py를 subprocess로 실행하고 출력에서 GCR 결과를 파싱
    
    Returns:
        {
            "physical_guard": [{"task": str, "gcr": float, ...}, ...],
            "baseline": [{"task": str, "gcr": float, ...}, ...]
        }
    """
    import re
    
    script_dir = Path(__file__).parent
    evaluate_script = script_dir / "evaluate_results.py"
    
    cmd = [
        sys.executable,
        str(evaluate_script),
        "--fp-number", str(fp_number),
        "--folder", folder,
        "--test_file", test_file
    ]
    
    physical_guard_results = []
    baseline_results = []
    
    try:
        process = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            universal_newlines=True,
            bufsize=1,
            cwd=script_dir.parent
        )
        
        in_comparison_table = False
        comparison_header_found = False
        
        # 출력 파싱
        for line in process.stdout:
            line = line.rstrip()
            print(line)  # 실시간 출력
            
            # SIDE-BY-SIDE COMPARISON 섹션 시작 감지
            if "SIDE-BY-SIDE COMPARISON" in line:
                in_comparison_table = True
                comparison_header_found = False
                continue
            
            # 비교 테이블 종료
            if in_comparison_table and line.startswith("=" * 70):
                if comparison_header_found:
                    in_comparison_table = False
                continue
            
            # 비교 테이블에서 데이터 파싱
            if in_comparison_table:
                # 헤더 라인 건너뛰기
                if "Task" in line and "Physical Guard" in line and "Baseline" in line:
                    comparison_header_found = True
                    continue
                
                if comparison_header_found and line and not line.startswith("-"):
                    # 형식: "Task name                   100.0%               100.0%"
                    # 또는: "Task name                   🏆 100.0%             0.0%"
                    # 정규식으로 마지막 두 개의 숫자.숫자% 패턴을 찾음
                    
                    # 🏆 제거
                    clean_line = line.replace('🏆', '').strip()
                    
                    # 모든 GCR 값 찾기 (숫자.숫자% 형식)
                    gcr_pattern = r'([\d.]+)%'
                    gcr_matches = re.findall(gcr_pattern, clean_line)
                    
                    if len(gcr_matches) >= 2:
                        # 마지막 두 개가 Physical Guard와 Baseline GCR
                        pg_gcr = float(gcr_matches[-2])
                        bl_gcr = float(gcr_matches[-1])
                        
                        # Task 이름: 마지막 두 개의 GCR 패턴을 제거한 나머지
                        # 정규식으로 마지막 두 개의 "숫자.숫자%" 패턴 제거
                        task_line = clean_line
                        # 뒤에서부터 두 개의 GCR 패턴 제거
                        for _ in range(2):
                            task_line = re.sub(r'\s*[\d.]+\%\s*$', '', task_line)
                        
                        task_name = task_line.strip()
                        
                        # 결과 저장
                        if task_name and pg_gcr is not None:
                            result = {
                                "task": task_name,
                                "method": "Physical Guard",
                                "status": "UNKNOWN",
                                "gcr": pg_gcr
                            }
                            # 중복 체크
                            existing = next((r for r in physical_guard_results if r["task"] == task_name), None)
                            if not existing:
                                physical_guard_results.append(result)
                            else:
                                existing.update(result)
                        
                        if task_name and bl_gcr is not None:
                            result = {
                                "task": task_name,
                                "method": "Baseline",
                                "status": "UNKNOWN",
                                "gcr": bl_gcr
                            }
                            # 중복 체크
                            existing = next((r for r in baseline_results if r["task"] == task_name), None)
                            if not existing:
                                baseline_results.append(result)
                            else:
                                existing.update(result)
            
            # 개별 결과도 파싱 (SIDE-BY-SIDE COMPARISON이 없는 경우 대비)
            if "[Physical Guard]" in line or "[Baseline]" in line:
                method = "Physical Guard" if "[Physical Guard]" in line else "Baseline"
                gcr_match = re.search(r'GCR:\s*([\d.]+)%', line)
                status_match = re.search(r'Task Result:\s*(\w+)', line)
                
                if gcr_match:
                    # 이전 줄에서 task 이름 찾기
                    # (이 부분은 개별 출력 형식에 따라 조정 필요)
                    pass
        
        process.wait()
        
        # Average GCR 추출 (출력에서 다시 파싱)
        # 이미 파싱한 결과를 사용하므로 추가 파싱 불필요
        
    except Exception as e:
        print(f"❌ evaluate_results.py 실행 실패: {e}")
        import traceback
        traceback.print_exc()
    
    return {
        "physical_guard": physical_guard_results,
        "baseline": baseline_results
    }

def run_baseline_subprocess(fp_number: int, task_file: str, folder: str = None) -> bool:
    """Baseline(ProgPrompt).py를 subprocess로 실행 (실시간 출력).
    folder를 주면 결과를 results/{folder}/FP{fp_number}/ 에 저장합니다."""
    script_dir = Path(__file__).parent
    project_root = script_dir.parent
    baseline_script = script_dir / "Baseline(ProgPrompt).py"
    
    if not baseline_script.exists():
        print(f"⚠️  Baseline 스크립트를 찾을 수 없습니다: {baseline_script}")
        return False
    
    cmd = [sys.executable, str(baseline_script), "--scene-number", str(fp_number)]
    if folder:
        out_dir = project_root / "results" / folder / f"FP{fp_number}"
        out_dir.mkdir(parents=True, exist_ok=True)
        cmd.extend(["--output-dir", str(out_dir)])
    if task_file:
        cmd.extend(["--task-file", task_file])
    
    try:
        process = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            universal_newlines=True,
            bufsize=1,
            cwd=script_dir.parent
        )
        
        # 실시간으로 출력
        for line in process.stdout:
            line = line.rstrip()
            print(line)
        
        process.wait()
        return process.returncode == 0
    except Exception as e:
        print(f"❌ Baseline 실행 실패: {e}")
        return False

def run_plan_generation(
    fp_number: int,
    folder: str,
    test_file: str,
    run_number: int
) -> bool:
    """
    Plan 생성만 실행
    - physical_guard.py 실행
    - Baseline(ProgPrompt).py 실행
    
    Returns:
        성공 여부
    """
    print(f"\n{'='*70}")
    print(f"📝 Run {run_number}/10 - FloorPlan{fp_number} (Plan 생성)")
    print(f"{'='*70}")
    
    # Step 1: physical_guard.py 실행 (plan 생성) → results/{folder}/FP{fp_number}/ 에 저장
    print(f"\n📝 Step 1: physical_guard.py 실행 중 (plan 생성)...")
    pg_success = run_physical_guard_subprocess(fp_number, test_file, folder)
    if not pg_success:
        print(f"⚠️  physical_guard.py 실행 실패")
        return False
    
    # Step 2: Baseline 실행 (plan 생성) → results/{folder}/FP{fp_number}/ 에 저장
    print(f"\n📝 Step 2: Baseline(ProgPrompt).py 실행 중 (plan 생성)...")
    bl_success = run_baseline_subprocess(fp_number, test_file, folder)
    if not bl_success:
        print(f"⚠️  Baseline 실행 실패")
        return False
    
    print(f"✅ Plan 생성 완료")
    return True

def run_evaluation_only(
    fp_number: int,
    folder: str,
    test_file: str,
    run_number: int
) -> Dict[str, Any]:
    """
    평가만 실행 (이미 생성된 plan 사용)
    - evaluate_results.py 실행하여 평가 및 GCR 추출
    
    Returns:
        {
            "physical_guard": [{"task": str, "gcr": float, ...}, ...],
            "baseline": [{"task": str, "gcr": float, ...}, ...]
        }
    """
    print(f"\n{'='*70}")
    print(f"📊 Run {run_number}/10 - FloorPlan{fp_number} (평가)")
    print(f"{'='*70}")
    
    # 결과 파일 찾기
    print(f"\n📝 생성된 plan 파일 찾기...")
    physical_guard_file, baseline_file = find_result_files(fp_number, folder)
    
    if not physical_guard_file and not baseline_file:
        print(f"⚠️  결과 파일을 찾을 수 없습니다.")
        return {"physical_guard": [], "baseline": []}
    
    if physical_guard_file:
        print(f"📖 Physical Guard 계획 파일: {physical_guard_file}")
    if baseline_file:
        print(f"📖 Baseline 계획 파일: {baseline_file}")
    
    # evaluate_results.py 실행하여 평가 및 GCR 추출
    print(f"\n📝 evaluate_results.py 실행 중 (평가 및 GCR 계산)...")
    results = run_evaluate_results_subprocess(fp_number, folder, test_file)
    
    return results

def run_single_evaluation(
    fp_number: int,
    folder: str,
    test_file: str,
    run_number: int
) -> Dict[str, Any]:
    """
    Plan 생성 + 평가 실행 (전체 프로세스)
    - physical_guard.py 실행
    - Baseline(ProgPrompt).py 실행
    - evaluate_results.py 로직으로 평가
    
    Returns:
        {
            "physical_guard": [{"task": str, "gcr": float, ...}, ...],
            "baseline": [{"task": str, "gcr": float, ...}, ...]
        }
    """
    print(f"\n{'='*70}")
    print(f"🔄 Run {run_number}/10 - FloorPlan{fp_number}")
    print(f"{'='*70}")
    
    # 이 run 시작 전 시간 기록 (이 시간 이후에 생성된 파일만 찾기)
    run_start_time = datetime.now()
    
    # Step 1: physical_guard.py 실행 (plan 생성) → results/{folder}/FP{fp_number}/ 에 저장
    print(f"\n📝 Step 1: physical_guard.py 실행 중 (plan 생성)...")
    pg_success = run_physical_guard_subprocess(fp_number, test_file, folder)
    if not pg_success:
        print(f"⚠️  physical_guard.py 실행 실패")
    
    # Step 2: Baseline 실행 (plan 생성) → results/{folder}/FP{fp_number}/ 에 저장
    print(f"\n📝 Step 2: Baseline(ProgPrompt).py 실행 중 (plan 생성)...")
    bl_success = run_baseline_subprocess(fp_number, test_file, folder)
    if not bl_success:
        print(f"⚠️  Baseline 실행 실패")
    
    # Step 3: 이 run에서 생성된 결과 파일 찾기 (results/{folder}/FP{fp_number}/ 우선) (파일 생성 대기)
    print(f"\n📝 Step 3: 생성된 plan 파일 찾기...")
    
    # 파일이 생성될 때까지 최대 10초 대기
    import time
    max_wait = 10
    wait_interval = 0.5
    waited = 0
    
    physical_guard_file = None
    baseline_file = None
    
    while waited < max_wait:
        physical_guard_file, baseline_file = find_result_files(fp_number, folder, after_time=run_start_time)
        if (physical_guard_file or baseline_file) or waited >= max_wait - wait_interval:
            break
        time.sleep(wait_interval)
        waited += wait_interval
    
    if not physical_guard_file and not baseline_file:
        print(f"⚠️  결과 파일을 찾을 수 없습니다. (대기 시간: {waited:.1f}초)")
        return {"physical_guard": [], "baseline": []}
    
    # Step 4: evaluate_results.py 실행하여 평가 및 GCR 추출
    print(f"\n📝 Step 4: evaluate_results.py 실행 중 (평가 및 GCR 계산)...")
    results = run_evaluate_results_subprocess(fp_number, folder, test_file)
    
    return results

def calculate_averages(all_results: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    모든 실행 결과의 평균을 계산
    
    Args:
        all_results: 각 실행의 결과 리스트
        
    Returns:
        {
            "physical_guard": {
                "task_name": {"avg_gcr": float, "gcr_list": [float, ...], ...},
                ...
            },
            "baseline": {...},
            "scene_avg": {
                "physical_guard": float,
                "baseline": float
            }
        }
    """
    pg_task_results = defaultdict(list)
    bl_task_results = defaultdict(list)
    
    # 각 실행 결과를 task별로 그룹화
    for run_result in all_results:
        for pg_res in run_result.get("physical_guard", []):
            task_name = pg_res.get("task", "Unknown")
            if "gcr" in pg_res:
                pg_task_results[task_name].append(pg_res["gcr"])
        
        for bl_res in run_result.get("baseline", []):
            task_name = bl_res.get("task", "Unknown")
            if "gcr" in bl_res:
                bl_task_results[task_name].append(bl_res["gcr"])
    
    # 평균 계산
    pg_avg = {}
    for task_name, gcr_list in pg_task_results.items():
        pg_avg[task_name] = {
            "avg_gcr": sum(gcr_list) / len(gcr_list),
            "gcr_list": gcr_list,
            "count": len(gcr_list)
        }
    
    bl_avg = {}
    for task_name, gcr_list in bl_task_results.items():
        bl_avg[task_name] = {
            "avg_gcr": sum(gcr_list) / len(gcr_list),
            "gcr_list": gcr_list,
            "count": len(gcr_list)
        }
    
    # Scene 평균 계산
    all_pg_gcr = [r["avg_gcr"] for r in pg_avg.values()]
    all_bl_gcr = [r["avg_gcr"] for r in bl_avg.values()]
    
    scene_avg = {
        "physical_guard": sum(all_pg_gcr) / len(all_pg_gcr) if all_pg_gcr else 0.0,
        "baseline": sum(all_bl_gcr) / len(all_bl_gcr) if all_bl_gcr else 0.0
    }
    
    return {
        "physical_guard": pg_avg,
        "baseline": bl_avg,
        "scene_avg": scene_avg
    }

def save_to_excel(all_scene_results: Dict[str, Dict[str, Any]], all_run_results: Dict[str, List[Dict[str, Any]]], output_file: str):
    """
    모든 Scene 결과를 Excel 파일로 저장
    
    Args:
        all_scene_results: 평균 결과
        all_run_results: 각 run의 상세 결과
        output_file: 출력 파일 경로
    """
    wb = openpyxl.Workbook()
    
    # 스타일 정의
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # 각 run별 상세 결과 시트 생성
    for fp_name, run_list in all_run_results.items():
        for run_data in run_list:
            run_num = run_data["run_number"]
            result = run_data["result"]
            
            # 시트 이름 생성 (Excel 시트 이름 제한: 31자)
            sheet_name = f"{fp_name}_R{run_num}"[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            # 헤더 작성
            headers = ["Task", "Physical Guard GCR (%)", "Baseline GCR (%)", "차이 (PG - BL)"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            
            # 데이터 작성
            pg_results = {r["task"]: r for r in result.get("physical_guard", [])}
            bl_results = {r["task"]: r for r in result.get("baseline", [])}
            
            all_tasks = set(pg_results.keys()) | set(bl_results.keys())
            row = 2
            
            for task_name in sorted(all_tasks):
                pg_res = pg_results.get(task_name, {})
                bl_res = bl_results.get(task_name, {})
                
                pg_gcr = pg_res.get("gcr", 0.0) if pg_res else None
                bl_gcr = bl_res.get("gcr", 0.0) if bl_res else None
                
                pg_gcr_val = pg_gcr if pg_gcr is not None else 0.0
                bl_gcr_val = bl_gcr if bl_gcr is not None else 0.0
                diff = pg_gcr_val - bl_gcr_val
                
                ws.cell(row=row, column=1, value=task_name)
                ws.cell(row=row, column=2, value=round(pg_gcr_val, 2) if pg_gcr is not None else "N/A")
                ws.cell(row=row, column=3, value=round(bl_gcr_val, 2) if bl_gcr is not None else "N/A")
                ws.cell(row=row, column=4, value=round(diff, 2) if (pg_gcr is not None and bl_gcr is not None) else "N/A")
                
                row += 1
            
            # 열 너비 조정
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
    
    # 각 Scene별 평균 시트 생성
    for fp_name, scene_data in all_scene_results.items():
        ws = wb.create_sheet(title=f"{fp_name}_Avg"[:31])
        
        # 헤더 작성
        headers = [
            "Task",
            "Physical Guard 평균 GCR (%)",
            "Baseline 평균 GCR (%)",
            "차이 (PG - BL)",
            "Physical Guard GCR (10회)",
            "Baseline GCR (10회)"
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        # 데이터 작성
        pg_data = scene_data["physical_guard"]
        bl_data = scene_data["baseline"]
        
        all_tasks = set(pg_data.keys()) | set(bl_data.keys())
        row = 2
        
        for task_name in sorted(all_tasks):
            pg_info = pg_data.get(task_name, {})
            bl_info = bl_data.get(task_name, {})
            
            pg_avg = pg_info.get("avg_gcr", 0.0)
            bl_avg = bl_info.get("avg_gcr", 0.0)
            diff = pg_avg - bl_avg
            
            pg_list = pg_info.get("gcr_list", [])
            bl_list = bl_info.get("gcr_list", [])
            
            # GCR 리스트를 문자열로 변환
            pg_list_str = ", ".join([f"{g:.1f}" for g in pg_list]) if pg_list else "N/A"
            bl_list_str = ", ".join([f"{g:.1f}" for g in bl_list]) if bl_list else "N/A"
            
            ws.cell(row=row, column=1, value=task_name)
            ws.cell(row=row, column=2, value=round(pg_avg, 2))
            ws.cell(row=row, column=3, value=round(bl_avg, 2))
            ws.cell(row=row, column=4, value=round(diff, 2))
            ws.cell(row=row, column=5, value=pg_list_str)
            ws.cell(row=row, column=6, value=bl_list_str)
            
            row += 1
        
        # Scene 평균 행 추가
        row += 1
        scene_avg = scene_data["scene_avg"]
        ws.cell(row=row, column=1, value="Scene 평균")
        ws.cell(row=row, column=2, value=round(scene_avg["physical_guard"], 2))
        ws.cell(row=row, column=3, value=round(scene_avg["baseline"], 2))
        ws.cell(row=row, column=4, value=round(scene_avg["physical_guard"] - scene_avg["baseline"], 2))
        
        # Scene 평균 행 스타일
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
    
    # 요약 시트 생성
    summary_ws = wb.create_sheet(title="Summary", index=0)
    
    summary_headers = [
        "Scene",
        "Physical Guard 평균 GCR (%)",
        "Baseline 평균 GCR (%)",
        "차이 (PG - BL)",
        "개선율 (%)"
    ]
    
    for col_idx, header in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    row = 2
    for fp_name in sorted(all_scene_results.keys()):
        scene_avg = all_scene_results[fp_name]["scene_avg"]
        pg_avg = scene_avg["physical_guard"]
        bl_avg = scene_avg["baseline"]
        diff = pg_avg - bl_avg
        improvement = ((pg_avg - bl_avg) / bl_avg * 100) if bl_avg > 0 else 0.0
        
        summary_ws.cell(row=row, column=1, value=fp_name)
        summary_ws.cell(row=row, column=2, value=round(pg_avg, 2))
        summary_ws.cell(row=row, column=3, value=round(bl_avg, 2))
        summary_ws.cell(row=row, column=4, value=round(diff, 2))
        summary_ws.cell(row=row, column=5, value=round(improvement, 2))
        row += 1
    
    # 전체 평균
    row += 1
    total_pg_avg = sum(r["scene_avg"]["physical_guard"] for r in all_scene_results.values()) / len(all_scene_results)
    total_bl_avg = sum(r["scene_avg"]["baseline"] for r in all_scene_results.values()) / len(all_scene_results)
    total_diff = total_pg_avg - total_bl_avg
    total_improvement = ((total_pg_avg - total_bl_avg) / total_bl_avg * 100) if total_bl_avg > 0 else 0.0
    
    summary_ws.cell(row=row, column=1, value="전체 평균")
    summary_ws.cell(row=row, column=2, value=round(total_pg_avg, 2))
    summary_ws.cell(row=row, column=3, value=round(total_bl_avg, 2))
    summary_ws.cell(row=row, column=4, value=round(total_diff, 2))
    summary_ws.cell(row=row, column=5, value=round(total_improvement, 2))
    
    # 전체 평균 행 스타일
    for col in range(1, 6):
        cell = summary_ws.cell(row=row, column=col)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    
    # 열 너비 조정
    for col in range(1, 6):
        summary_ws.column_dimensions[get_column_letter(col)].width = 25
    
    # 기본 시트 제거
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # 파일 저장
    wb.save(output_file)
    print(f"\n✅ Excel 파일 저장 완료: {output_file}")

def format_time(seconds: float) -> str:
    """초를 읽기 쉬운 시간 형식으로 변환"""
    if seconds < 60:
        return f"{int(seconds)}초"
    elif seconds < 3600:
        minutes = int(seconds // 60)
        secs = int(seconds % 60)
        return f"{minutes}분 {secs}초"
    else:
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        secs = int(seconds % 60)
        return f"{hours}시간 {minutes}분 {secs}초"

def print_progress(current: int, total: int, start_time: datetime, current_task: str = ""):
    """진행 상황 출력"""
    elapsed = (datetime.now() - start_time).total_seconds()
    progress = (current / total) * 100
    
    if current > 0:
        avg_time_per_task = elapsed / current
        remaining_tasks = total - current
        estimated_remaining = avg_time_per_task * remaining_tasks
        estimated_total = elapsed + estimated_remaining
    else:
        estimated_remaining = 0
        estimated_total = 0
    
    print(f"\n{'='*70}")
    print(f"📊 진행 상황: {current}/{total} ({progress:.1f}%)")
    print(f"⏱️  경과 시간: {format_time(elapsed)}")
    if estimated_remaining > 0:
        print(f"⏳ 예상 남은 시간: {format_time(estimated_remaining)}")
        print(f"🕐 예상 완료 시간: {format_time(estimated_total)}")
    if current_task:
        print(f"🔄 현재 작업: {current_task}")
    print(f"{'='*70}\n")
    
    # 진행 바 표시
    bar_length = 50
    filled = int(bar_length * current / total)
    bar = "█" * filled + "░" * (bar_length - filled)
    print(f"[{bar}] {progress:.1f}%")

def main():
    """메인 함수"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Batch Evaluation Script - 각 FloorPlan을 10번씩 실행하여 평균 GCR 계산"
    )
    parser.add_argument(
        "--mode",
        type=str,
        choices=["full", "plan-only", "evaluate-only"],
        default="full",
        help="실행 모드: full (전체), plan-only (plan 생성만), evaluate-only (평가만)"
    )
    parser.add_argument(
        "--fp-number",
        type=int,
        default=None,
        help="평가할 FloorPlan 번호 (예: 1, 216, 325, 403). 지정하지 않으면 모든 FloorPlan 실행"
    )
    
    args = parser.parse_args()
    
    script_dir = Path(__file__).parent
    project_root = script_dir.parent
    
    print("\n" + "="*70)
    print("🔄 Batch Evaluation Script")
    if args.mode == "plan-only":
        print("   모드: Plan 생성만")
    elif args.mode == "evaluate-only":
        print("   모드: 평가만 (이미 생성된 plan 사용)")
    else:
        print("   모드: 전체 (Plan 생성 + 평가)")
    print("="*70)
    
    # --fp-number 지정 시: 해당 FP의 Task 목록을 보여주고 번호로 선택 → 그 Task만 10번 실행
    selected_task_index = None
    single_task_path = None
    
    if args.fp_number:
        floor_plans = [args.fp_number]
        task_list = load_tasks_from_fp(project_root, args.fp_number)
        if not task_list:
            print(f"❌ FloorPlan{args.fp_number}의 Task 파일을 읽을 수 없거나 비어 있습니다.")
            sys.exit(1)
        n_tasks = len(task_list)
        print(f"\n📋 FloorPlan{args.fp_number} 의 Task 목록 (총 {n_tasks}개):")
        for i, t in enumerate(task_list, 1):
            task_str = t.get("task", "(task 없음)")
            print(f"   {i}. {task_str}")
        while True:
            try:
                sel = input(f"\n실행할 Task 번호를 선택하세요 (1-{n_tasks}): ").strip()
                task_index = int(sel)
                if 1 <= task_index <= n_tasks:
                    break
            except ValueError:
                pass
            print(f"   잘못된 입력입니다. 1~{n_tasks} 사이 숫자를 입력하세요.")
        selected_task_def = task_list[task_index - 1]
        single_task_path = project_root / "data" / "final_test" / f"_single_task_fp{args.fp_number}_task{task_index}.json"
        with open(single_task_path, "w", encoding="utf-8") as f:
            json.dump([selected_task_def], f, indent=2, ensure_ascii=False)
        selected_task_index = task_index
        print(f"   선택: Task {task_index} — 10번 실행 후 결과 저장")
    else:
        floor_plans = [1, 216, 325, 403]
        print(f"📋 기본 FloorPlan 목록: {floor_plans}")
    
    num_runs = 10
    
    # 전체 작업 수 계산
    total_tasks = len(floor_plans) * num_runs
    current_task_num = 0
    start_time = datetime.now()
    
    all_scene_results = {}
    all_run_results = {}  # 각 run의 상세 결과 저장
    
    try:
        for fp_idx, fp_number in enumerate(floor_plans, 1):
            print(f"\n{'#'*70}")
            print(f"# FloorPlan{fp_number} 시작 ({fp_idx}/{len(floor_plans)})")
            print(f"{'#'*70}")
            
            folder = get_folder_from_fp_number(fp_number)
            # 단일 Task 선택 시 해당 task만 담긴 JSON 사용, 아니면 전체 FloorPlan JSON 사용
            if args.fp_number and selected_task_index is not None and single_task_path is not None:
                test_file = single_task_path
            else:
                test_file = project_root / "data" / "final_test" / f"FloorPlan{fp_number}.json"
            test_file = Path(test_file)
            
            if not test_file.exists():
                print(f"⚠️  테스트 파일을 찾을 수 없습니다: {test_file}")
                continue
            test_file_str = str(test_file)
            
            fp_name = f"FloorPlan{fp_number}"
            all_run_results[fp_name] = []
            
            # 10번 실행
            all_results = []
            for run_num in range(1, num_runs + 1):
                current_task_num += 1
                current_task = f"FloorPlan{fp_number} - Run {run_num}/{num_runs}"
                
                # 진행 상황 출력
                print_progress(current_task_num, total_tasks, start_time, current_task)
                
                try:
                    if args.mode == "plan-only":
                        # Plan 생성만
                        success = run_plan_generation(
                            fp_number=fp_number,
                            folder=folder,
                            test_file=test_file_str,
                            run_number=run_num
                        )
                        if success:
                            print(f"✅ {current_task} 완료")
                        else:
                            print(f"❌ {current_task} 실패")
                    
                    elif args.mode == "evaluate-only":
                        # 평가만
                        result = run_evaluation_only(
                            fp_number=fp_number,
                            folder=folder,
                            test_file=test_file_str,
                            run_number=run_num
                        )
                        all_results.append(result)
                        # 각 run의 결과 저장
                        all_run_results[fp_name].append({
                            "run_number": run_num,
                            "result": result
                        })
                        print(f"✅ {current_task} 완료")
                    
                    else:  # full
                        # Plan 생성 + 평가
                        result = run_single_evaluation(
                            fp_number=fp_number,
                            folder=folder,
                            test_file=test_file_str,
                            run_number=run_num
                        )
                        all_results.append(result)
                        # 각 run의 결과 저장
                        all_run_results[fp_name].append({
                            "run_number": run_num,
                            "result": result
                        })
                        print(f"✅ {current_task} 완료")
                
                except KeyboardInterrupt:
                    # Ctrl+C가 내부에서 발생한 경우
                    raise
                except Exception as e:
                    print(f"❌ {current_task} 실패: {e}")
                    import traceback
                    traceback.print_exc()
                    continue
            
            # 평가 모드인 경우에만 평균 계산
            if args.mode != "plan-only" and all_results:
                # 평균 계산
                avg_results = calculate_averages(all_results)
                all_scene_results[fp_name] = avg_results
                
                print(f"\n✅ FloorPlan{fp_number} 완료 ({fp_idx}/{len(floor_plans)})")
                print(f"   Physical Guard 평균 GCR: {avg_results['scene_avg']['physical_guard']:.2f}%")
                print(f"   Baseline 평균 GCR: {avg_results['scene_avg']['baseline']:.2f}%")
            elif args.mode == "plan-only":
                print(f"\n✅ FloorPlan{fp_number} Plan 생성 완료 ({fp_idx}/{len(floor_plans)})")
        
        # 평가 모드인 경우에만 Excel 파일로 저장
        if args.mode != "plan-only":
            if all_scene_results or all_run_results:
                print(f"\n💾 Excel 파일 저장 중...")
                # 단일 Task 선택 시: FP{num}_task{index}_result.xlsx
                # 단일 FloorPlan만 지정 시: FP{num}_result.xlsx
                if args.fp_number and selected_task_index is not None:
                    output_file = project_root / f"FP{args.fp_number}_task{selected_task_index}_result.xlsx"
                elif args.fp_number:
                    output_file = project_root / f"FP{args.fp_number}_result.xlsx"
                else:
                    output_file = project_root / "result_half_test.xlsx"
                save_to_excel(all_scene_results, all_run_results, str(output_file))
        
        # 단일 Task용 임시 JSON 삭제
        if single_task_path is not None and Path(single_task_path).exists():
            try:
                Path(single_task_path).unlink()
            except Exception:
                pass
        
        # 최종 진행 상황
        total_elapsed = (datetime.now() - start_time).total_seconds()
        print_progress(total_tasks, total_tasks, start_time, "모든 작업 완료")
        
        print("\n" + "="*70)
        if args.mode == "plan-only":
            print("✅ 모든 Plan 생성 완료!")
        else:
            print("✅ 모든 평가 완료!")
        print(f"⏱️  총 소요 시간: {format_time(total_elapsed)}")
        print("="*70)
    
    except KeyboardInterrupt:
        print("\n\n" + "="*70)
        print("⚠️  사용자에 의해 실행이 중단되었습니다 (Ctrl+C)")
        print("="*70)
        
        # 평가 모드인 경우에만 결과 저장
        mode = args.mode if 'args' in locals() else "full"
        fp_number = args.fp_number if 'args' in locals() and hasattr(args, 'fp_number') else None
        if mode != "plan-only":
            # 현재까지 수집된 결과가 있으면 저장
            if all_scene_results or all_run_results:
                print(f"\n💾 지금까지 실행된 결과를 Excel 파일로 저장 중...")
                if fp_number and selected_task_index is not None:
                    output_file = project_root / f"FP{fp_number}_task{selected_task_index}_result.xlsx"
                elif fp_number:
                    output_file = project_root / f"FP{fp_number}_result.xlsx"
                else:
                    output_file = project_root / "result_half_test.xlsx"
                
                # 완료되지 않은 FloorPlan의 평균 계산
                for fp_name, run_list in all_run_results.items():
                    if fp_name not in all_scene_results and run_list:
                        # 해당 FloorPlan의 결과만 평균 계산
                        results = [r["result"] for r in run_list]
                        if results:
                            avg_results = calculate_averages(results)
                            all_scene_results[fp_name] = avg_results
                
                try:
                    save_to_excel(all_scene_results, all_run_results, str(output_file))
                    
                    total_elapsed = (datetime.now() - start_time).total_seconds()
                    completed_runs = sum(len(runs) for runs in all_run_results.values())
                    
                    print(f"\n✅ 저장 완료!")
                    print(f"   완료된 Run 수: {completed_runs}/{total_tasks}")
                    print(f"   경과 시간: {format_time(total_elapsed)}")
                    print(f"   저장된 파일: {output_file}")
                except Exception as e:
                    print(f"\n❌ Excel 파일 저장 실패: {e}")
                    import traceback
                    traceback.print_exc()
            else:
                print("\n⚠️  저장할 결과가 없습니다.")
        
        # 단일 Task용 임시 JSON 삭제
        if single_task_path is not None and Path(single_task_path).exists():
            try:
                Path(single_task_path).unlink()
            except Exception:
                pass
        
        print("\n" + "="*70)
        sys.exit(0)

if __name__ == "__main__":
    main()
